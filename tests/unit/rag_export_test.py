"""Unit coverage for workbook, summary, CSV, and atomic export publication."""

from __future__ import annotations

import csv
from pathlib import Path
from tempfile import TemporaryDirectory
import unittest
from unittest.mock import patch

from openpyxl import load_workbook

import rag


def _runtime_row(
    *,
    question_id: str,
    category: str,
    question: str,
    answer: str = "",
    evidence: str = "",
    confidence: str = "",
    status: str = "",
    reviewer_notes: str = "",
) -> dict[str, object]:
    """Build one canonical runtime row for export-focused tests."""
    row = {
        "Question ID": question_id,
        "Category": category,
        "Question": question,
        "Answer": answer,
        "Evidence": evidence,
        "Confidence": confidence,
        "Status": status,
        "Reviewer Notes": reviewer_notes,
        "question_id": question_id,
        "category": category,
        "question": question,
        **rag.make_result_row_defaults(),
    }
    row["answer"] = answer
    row["confidence_band"] = confidence
    row["status"] = status
    row["reviewer_note"] = reviewer_notes
    row["evidence_labels"] = evidence.split("; ") if evidence else []
    return row


def _runtime_questionnaire(*rows: dict[str, object]) -> rag.RuntimeQuestionnaire:
    """Build one lightweight questionnaire fixture in workbook order."""
    return rag.RuntimeQuestionnaire(
        workbook_path=Path("data/questionnaires/Demo_Security_Questionnaire.xlsx"),
        visible_columns=rag.VISIBLE_EXPORT_COLUMNS,
        rows=[dict(row) for row in rows],
    )


def _citation(
    *,
    chunk_id: str,
    label: str,
    snippet: str,
    source: str,
    section: str | None = None,
    page: int | None = None,
) -> rag.ResolvedEvidenceCitation:
    """Build one reviewer-facing citation fixture."""
    return rag.ResolvedEvidenceCitation(
        chunk_id=chunk_id,
        display_label=label,
        snippet_text=snippet,
        source=source,
        source_path=rag.RUNTIME_EVIDENCE_DIR / source,
        doc_type=(
            rag.DOCUMENT_TYPE_PDF
            if source.endswith(".pdf")
            else rag.DOCUMENT_TYPE_POLICY
        ),
        section=section,
        page=page,
    )


def _answer_result(
    *,
    answer: str,
    answer_type: str,
    citation_ids: tuple[str, ...],
    citations: tuple[rag.ResolvedEvidenceCitation, ...],
    confidence_score: float,
    confidence_band: str,
    status: str,
    reviewer_note: str,
) -> rag.GeneratedAnswerResult:
    """Build one deterministic answer result fixture."""
    return rag.GeneratedAnswerResult(
        answer=answer,
        answer_type=answer_type,
        citation_ids=citation_ids,
        citations=citations,
        confidence_score=confidence_score,
        confidence_band=confidence_band,
        status=status,
        reviewer_note=reviewer_note,
    )


class RagExportTest(unittest.TestCase):
    """Verify export artifacts stay stable, readable, and coherently published."""

    def test_write_answered_questionnaire_preserves_order_labels_and_styling(self):
        """Workbook exports should keep row order, friendly evidence labels, and styling."""
        questionnaire = _runtime_questionnaire(
            _runtime_row(
                question_id="Q01",
                category="Encryption",
                question="Is customer data encrypted at rest?",
            ),
            _runtime_row(
                question_id="Q02",
                category="Residency",
                question="Can customers choose their data region?",
            ),
        )
        questionnaire.rows[0] = rag.update_row_with_answer_result(
            questionnaire.rows[0],
            _answer_result(
                answer="Yes. Customer data is encrypted at rest.",
                answer_type=rag.ANSWER_TYPE_SUPPORTED,
                citation_ids=("enc_001", "soc2_001"),
                citations=(
                    _citation(
                        chunk_id="enc_001",
                        label="Encryption Policy — Data at Rest",
                        snippet="Customer data is encrypted at rest with AES-256.",
                        source=rag.ENCRYPTION_POLICY_FILE_NAME,
                        section="Data at Rest",
                    ),
                    _citation(
                        chunk_id="soc2_001",
                        label="SOC 2 Summary — Page 2",
                        snippet="The SOC 2 audit describes encryption controls.",
                        source=rag.SOC2_SUMMARY_FILE_NAME,
                        page=2,
                    ),
                ),
                confidence_score=rag.SUPPORTED_WITH_TWO_PLUS_CITATIONS_SCORE,
                confidence_band=rag.CONFIDENCE_BAND_HIGH,
                status=rag.STATUS_READY_FOR_REVIEW,
                reviewer_note="",
            ),
            index_action=rag.INDEX_ACTION_REUSED,
            run_id="run-001",
        )
        questionnaire.rows[1] = rag.update_row_with_answer_result(
            questionnaire.rows[1],
            _answer_result(
                answer="Not stated.",
                answer_type=rag.ANSWER_TYPE_UNSUPPORTED,
                citation_ids=(),
                citations=(),
                confidence_score=rag.UNSUPPORTED_SCORE,
                confidence_band=rag.CONFIDENCE_BAND_LOW,
                status=rag.STATUS_NEEDS_REVIEW,
                reviewer_note="Region selection is not described in the bundled evidence.",
            ),
            index_action=rag.INDEX_ACTION_REUSED,
            run_id="run-001",
        )

        with TemporaryDirectory() as tmpdir:
            output_path = rag.write_answered_questionnaire(
                questionnaire,
                output_dir=Path(tmpdir),
            )
            workbook = load_workbook(output_path)
            worksheet = workbook[rag.QUESTION_SHEET_NAME]
            rows = list(worksheet.iter_rows(values_only=True))

            self.assertEqual(rows[0], tuple(rag.VISIBLE_EXPORT_COLUMNS))
            self.assertEqual(rows[1][0], "Q01")
            self.assertEqual(
                rows[1][4],
                "Encryption Policy — Data at Rest; SOC 2 Summary — Page 2",
            )
            self.assertEqual(rows[2][0], "Q02")
            self.assertEqual(rows[2][6], rag.STATUS_NEEDS_REVIEW)
            self.assertIn(rows[1][7], ("", None))
            self.assertEqual(worksheet.freeze_panes, "A2")
            self.assertEqual(worksheet.auto_filter.ref, "A1:H3")
            self.assertTrue(all(cell.font.bold for cell in worksheet[1]))
            self.assertTrue(worksheet["C2"].alignment.wrap_text)
            self.assertTrue(worksheet["D2"].alignment.wrap_text)
            self.assertEqual(
                worksheet["G2"].fill.fgColor.rgb,
                rag.STATUS_FILL_RGB_BY_NAME[rag.READY_STATUS_FILL],
            )
            self.assertEqual(
                worksheet["G3"].fill.fgColor.rgb,
                rag.STATUS_FILL_RGB_BY_NAME[rag.REVIEW_STATUS_FILL],
            )
            workbook.close()

    def test_write_review_summary_and_needs_review_csv_use_priority_order_and_provenance(self):
        """Summary and CSV exports should agree on review ordering and reviewer notes."""
        questionnaire = _runtime_questionnaire(
            _runtime_row(
                question_id="Q01",
                category="Encryption",
                question="Is customer data encrypted at rest?",
            ),
            _runtime_row(
                question_id="Q02",
                category="Backups",
                question="Are backups performed daily?",
            ),
            _runtime_row(
                question_id="Q03",
                category="Residency",
                question="Can customers choose their data region?",
            ),
        )
        questionnaire.rows[0] = rag.update_row_with_answer_result(
            questionnaire.rows[0],
            _answer_result(
                answer="Yes.",
                answer_type=rag.ANSWER_TYPE_SUPPORTED,
                citation_ids=(),
                citations=(),
                confidence_score=rag.SUPPORTED_WITH_ONE_CITATION_SCORE,
                confidence_band=rag.CONFIDENCE_BAND_MEDIUM,
                status=rag.STATUS_READY_FOR_REVIEW,
                reviewer_note="",
            ),
            index_action=rag.INDEX_ACTION_REUSED,
            run_id="run-002",
        )
        questionnaire.rows[1] = rag.update_row_with_answer_result(
            questionnaire.rows[1],
            _answer_result(
                answer="Not stated.",
                answer_type=rag.ANSWER_TYPE_UNSUPPORTED,
                citation_ids=(),
                citations=(),
                confidence_score=0.20,
                confidence_band=rag.CONFIDENCE_BAND_LOW,
                status=rag.STATUS_NEEDS_REVIEW,
                reviewer_note="Backup cadence is not explicit in the bundled evidence.",
            ),
            index_action=rag.INDEX_ACTION_REUSED,
            run_id="run-002",
        )
        questionnaire.rows[2] = rag.update_row_with_answer_result(
            questionnaire.rows[2],
            _answer_result(
                answer="Partially stated.",
                answer_type=rag.ANSWER_TYPE_PARTIAL,
                citation_ids=(),
                citations=(),
                confidence_score=rag.PARTIAL_SCORE,
                confidence_band=rag.CONFIDENCE_BAND_MEDIUM,
                status=rag.STATUS_NEEDS_REVIEW,
                reviewer_note="Region selection details are incomplete.",
            ),
            index_action=rag.INDEX_ACTION_REUSED,
            run_id="run-002",
        )

        with TemporaryDirectory() as tmpdir:
            output_dir = Path(tmpdir)
            summary_path = rag.write_review_summary(
                questionnaire,
                output_dir=output_dir,
                completed_at="2026-04-27T22:04:00Z",
                workspace_hash="workspace-hash-002",
            )
            csv_path = rag.write_needs_review_csv(
                questionnaire,
                output_dir=output_dir,
            )
            summary_lines = summary_path.read_text(encoding="utf-8").splitlines()
            self.assertEqual(summary_lines[0], "Completed Run: 2026-04-27T22:04:00Z")
            self.assertEqual(summary_lines[1], "Workspace Hash: workspace-hash-002")
            self.assertEqual(summary_lines[2], "Index State: reused")
            self.assertIn("- Total Questions: 3", summary_lines)
            self.assertIn("- Ready for Review: 1", summary_lines)
            self.assertIn("- Needs Review: 2", summary_lines)
            self.assertLess(
                summary_lines.index(
                    "- Q02: Backup cadence is not explicit in the bundled evidence."
                ),
                summary_lines.index("- Q03: Region selection details are incomplete."),
            )

            with csv_path.open(newline="", encoding="utf-8") as handle:
                rows = list(csv.DictReader(handle))
            self.assertEqual(list(rows[0].keys()), list(rag.REVIEW_QUEUE_COLUMNS))
            self.assertEqual(
                [row["Question ID"] for row in rows],
                ["Q02", "Q03"],
            )
            self.assertEqual(
                [row["Reviewer Notes"] for row in rows],
                [
                    "Backup cadence is not explicit in the bundled evidence.",
                    "Region selection details are incomplete.",
                ],
            )

    def test_publish_export_packet_swaps_in_a_coherent_run_and_returns_paths(self):
        """Successful publish should replace the packet as a unit and preserve a backup."""
        questionnaire = _runtime_questionnaire(
            _runtime_row(
                question_id="Q01",
                category="Encryption",
                question="Is customer data encrypted at rest?",
            ),
            _runtime_row(
                question_id="Q02",
                category="Residency",
                question="Can customers choose their data region?",
            ),
        )
        questionnaire.rows[0] = rag.update_row_with_answer_result(
            questionnaire.rows[0],
            _answer_result(
                answer="Yes.",
                answer_type=rag.ANSWER_TYPE_SUPPORTED,
                citation_ids=(),
                citations=(),
                confidence_score=rag.SUPPORTED_WITH_ONE_CITATION_SCORE,
                confidence_band=rag.CONFIDENCE_BAND_MEDIUM,
                status=rag.STATUS_READY_FOR_REVIEW,
                reviewer_note="",
            ),
            index_action=rag.INDEX_ACTION_REUSED,
            run_id="run-003",
        )
        questionnaire.rows[1] = rag.update_row_with_answer_result(
            questionnaire.rows[1],
            _answer_result(
                answer="Not stated.",
                answer_type=rag.ANSWER_TYPE_UNSUPPORTED,
                citation_ids=(),
                citations=(),
                confidence_score=rag.UNSUPPORTED_SCORE,
                confidence_band=rag.CONFIDENCE_BAND_LOW,
                status=rag.STATUS_NEEDS_REVIEW,
                reviewer_note="Region selection is not described in the bundled evidence.",
            ),
            index_action=rag.INDEX_ACTION_REUSED,
            run_id="run-003",
        )

        with TemporaryDirectory() as tmpdir:
            output_dir = Path(tmpdir) / "outputs"
            output_dir.mkdir()
            (output_dir / rag.REVIEW_SUMMARY_FILE_NAME).write_text(
                "old summary",
                encoding="utf-8",
            )
            (output_dir / rag.NEEDS_REVIEW_FILE_NAME).write_text(
                "old csv",
                encoding="utf-8",
            )
            (output_dir / rag.ANSWERED_QUESTIONNAIRE_FILE_NAME).write_text(
                "old workbook placeholder",
                encoding="utf-8",
            )

            packet = rag.publish_export_packet(
                questionnaire,
                output_dir=output_dir,
                completed_at="2026-04-27T22:05:00Z",
                workspace_hash="workspace-hash-003",
            )

            self.assertEqual(packet.output_dir, output_dir)
            self.assertEqual(packet.run_id, "run-003")
            self.assertEqual(packet.completed_at, "2026-04-27T22:05:00Z")
            self.assertEqual(packet.workspace_hash, "workspace-hash-003")
            self.assertTrue(packet.answered_questionnaire_path.exists())
            self.assertTrue(packet.review_summary_path.exists())
            self.assertTrue(packet.needs_review_csv_path.exists())
            self.assertEqual(
                packet.review_summary_path.read_text(encoding="utf-8").splitlines()[0],
                "Completed Run: 2026-04-27T22:05:00Z",
            )

            backups = sorted(output_dir.parent.glob(".outputs-backup-run-003-*"))
            self.assertEqual(len(backups), 1)
            self.assertEqual(
                (backups[0] / rag.REVIEW_SUMMARY_FILE_NAME).read_text(encoding="utf-8"),
                "old summary",
            )

    def test_publish_export_packet_preserves_previous_packet_when_staging_fails(self):
        """A failed publish attempt should leave the previous good packet untouched."""
        questionnaire = _runtime_questionnaire(
            _runtime_row(
                question_id="Q01",
                category="Encryption",
                question="Is customer data encrypted at rest?",
                answer="Yes.",
                confidence=rag.CONFIDENCE_BAND_HIGH,
                status=rag.STATUS_READY_FOR_REVIEW,
            )
        )
        questionnaire.rows[0]["answer"] = "Yes."
        questionnaire.rows[0]["confidence_band"] = rag.CONFIDENCE_BAND_HIGH
        questionnaire.rows[0]["status"] = rag.STATUS_READY_FOR_REVIEW
        questionnaire.rows[0]["run_id"] = "run-004"
        questionnaire.rows[0]["index_action"] = rag.INDEX_ACTION_REUSED

        with TemporaryDirectory() as tmpdir:
            output_dir = Path(tmpdir) / "outputs"
            output_dir.mkdir()
            previous_summary = output_dir / rag.REVIEW_SUMMARY_FILE_NAME
            previous_summary.write_text("previous good packet", encoding="utf-8")

            with patch(
                "rag.write_needs_review_csv",
                side_effect=RuntimeError("simulated export failure"),
            ):
                with self.assertRaisesRegex(RuntimeError, "simulated export failure"):
                    rag.publish_export_packet(
                        questionnaire,
                        output_dir=output_dir,
                        completed_at="2026-04-27T22:06:00Z",
                        workspace_hash="workspace-hash-004",
                    )

            self.assertEqual(
                previous_summary.read_text(encoding="utf-8"),
                "previous good packet",
            )


if __name__ == "__main__":
    unittest.main()
