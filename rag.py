"""Core contracts and RAG helpers for the Security Questionnaire Agent."""

from __future__ import annotations

from dataclasses import dataclass, replace
import json
import os
from pathlib import Path
from shlex import join as shell_join
from typing import Any, Callable, Final, Mapping, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT: Final[Path] = Path(__file__).resolve().parent

APP_TITLE: Final[str] = "AI Security Questionnaire Copilot"
APP_SUBTITLE: Final[str] = (
    "Answer one curated security questionnaire from one bundled evidence pack."
)
DEMO_MODE_LABEL: Final[str] = (
    "Demo Mode: curated questionnaire + bundled evidence pack"
)

SEED_DATA_DIR: Final[Path] = REPO_ROOT / "seed_data"
SEED_QUESTIONNAIRE_DIR: Final[Path] = SEED_DATA_DIR / "questionnaire"
SEED_EVIDENCE_DIR: Final[Path] = SEED_DATA_DIR / "evidence"

DATA_DIR: Final[Path] = REPO_ROOT / "data"
RUNTIME_QUESTIONNAIRES_DIR: Final[Path] = DATA_DIR / "questionnaires"
RUNTIME_EVIDENCE_DIR: Final[Path] = DATA_DIR / "evidence"
OUTPUTS_DIR: Final[Path] = DATA_DIR / "outputs"
CHROMA_DIR: Final[Path] = DATA_DIR / "chroma"

QUESTIONNAIRE_FILE_NAME: Final[str] = "Demo_Security_Questionnaire.xlsx"
SOC2_SUMMARY_FILE_NAME: Final[str] = "AcmeCloud_SOC2_Summary.pdf"
ENCRYPTION_POLICY_FILE_NAME: Final[str] = "Encryption_Policy.md"
ACCESS_CONTROL_POLICY_FILE_NAME: Final[str] = "Access_Control_Policy.md"
INCIDENT_RESPONSE_POLICY_FILE_NAME: Final[str] = "Incident_Response_Policy.md"
BACKUP_RECOVERY_POLICY_FILE_NAME: Final[str] = "Backup_and_Recovery_Policy.md"
MANIFEST_FILE_NAME: Final[str] = "workspace_manifest.json"

ANSWERED_QUESTIONNAIRE_FILE_NAME: Final[str] = "Answered_Questionnaire.xlsx"
REVIEW_SUMMARY_FILE_NAME: Final[str] = "Review_Summary.md"
NEEDS_REVIEW_FILE_NAME: Final[str] = "Needs_Review.csv"

QUESTION_SHEET_NAME: Final[str] = "Questions"

SEED_QUESTION_COLUMNS: Final[tuple[str, ...]] = (
    "Question ID",
    "Category",
    "Question",
)
VISIBLE_OUTPUT_COLUMNS: Final[tuple[str, ...]] = (
    "Answer",
    "Evidence",
    "Confidence",
    "Status",
    "Reviewer Notes",
)
VISIBLE_EXPORT_COLUMNS: Final[tuple[str, ...]] = (
    *SEED_QUESTION_COLUMNS,
    *VISIBLE_OUTPUT_COLUMNS,
)
MAIN_RESULTS_TABLE_COLUMNS: Final[tuple[str, ...]] = (
    "Question ID",
    "Category",
    "Answer",
    "Confidence",
    "Status",
)
REVIEW_QUEUE_COLUMNS: Final[tuple[str, ...]] = (
    "Question ID",
    "Category",
    "Question",
    "Answer",
    "Confidence",
    "Status",
    "Reviewer Notes",
)
SUMMARY_CARD_LABELS: Final[tuple[str, ...]] = (
    "Questions",
    "Ready for Review",
    "Needs Review",
    "Sources Indexed",
)

STATUS_READY_FOR_REVIEW: Final[str] = "Ready for Review"
STATUS_NEEDS_REVIEW: Final[str] = "Needs Review"

ANSWER_TYPE_SUPPORTED: Final[str] = "supported"
ANSWER_TYPE_PARTIAL: Final[str] = "partial"
ANSWER_TYPE_UNSUPPORTED: Final[str] = "unsupported"
ANSWER_TYPES: Final[tuple[str, ...]] = (
    ANSWER_TYPE_SUPPORTED,
    ANSWER_TYPE_PARTIAL,
    ANSWER_TYPE_UNSUPPORTED,
)

CONFIDENCE_BAND_HIGH: Final[str] = "High"
CONFIDENCE_BAND_MEDIUM: Final[str] = "Medium"
CONFIDENCE_BAND_LOW: Final[str] = "Low"

READY_STATUS_FILL: Final[str] = "light_green"
REVIEW_STATUS_FILL: Final[str] = "light_amber"
ANSWERED_WORKBOOK_COLUMN_WIDTHS: Final[Mapping[str, float]] = {
    "Question ID": 14,
    "Category": 18,
    "Question": 52,
    "Answer": 52,
    "Evidence": 32,
    "Confidence": 14,
    "Status": 18,
    "Reviewer Notes": 40,
}
STATUS_FILL_RGB_BY_NAME: Final[Mapping[str, str]] = {
    READY_STATUS_FILL: "FFE2F0D9",
    REVIEW_STATUS_FILL: "FFFFF2CC",
}
WRAPPED_EXPORT_COLUMNS: Final[frozenset[str]] = frozenset(
    {"Question", "Answer", "Evidence", "Reviewer Notes"}
)

SUPPORTED_WITH_ONE_CITATION_SCORE: Final[float] = 0.78
SUPPORTED_WITH_TWO_PLUS_CITATIONS_SCORE: Final[float] = 0.90
PARTIAL_SCORE: Final[float] = 0.55
UNSUPPORTED_SCORE: Final[float] = 0.30
FAIL_CLOSED_SCORE: Final[float] = 0.25
FAIL_CLOSED_ANSWER: Final[str] = (
    "Not stated. The available evidence does not support a reliable answer."
)
HIGH_CONFIDENCE_THRESHOLD: Final[float] = 0.85
MEDIUM_CONFIDENCE_THRESHOLD: Final[float] = 0.60
REVIEW_QUEUE_THRESHOLD: Final[float] = 0.75

COLLECTION_NAME: Final[str] = "security_questionnaire_demo"
CHUNK_SIZE_CHARS: Final[int] = 700
CHUNK_OVERLAP_CHARS: Final[int] = 100
RETRIEVAL_TOP_K: Final[int] = 5
MAX_VISIBLE_CITATIONS: Final[int] = 2
MODEL_RETRY_LIMIT: Final[int] = 1
DEFAULT_OPENAI_ANSWER_MODEL: Final[str] = "gpt-4.1-mini"
ANSWER_OUTPUT_KEYS: Final[tuple[str, ...]] = (
    "answer",
    "answer_type",
    "citation_ids",
    "reviewer_note",
)
ANSWER_RESPONSE_SCHEMA_NAME: Final[str] = "security_questionnaire_answer"
ANSWER_RESPONSE_JSON_SCHEMA: Final[dict[str, object]] = {
    "type": "object",
    "additionalProperties": False,
    "properties": {
        "answer": {"type": "string"},
        "answer_type": {"type": "string", "enum": list(ANSWER_TYPES)},
        "citation_ids": {
            "type": "array",
            "items": {"type": "string"},
        },
        "reviewer_note": {"type": "string"},
    },
    "required": list(ANSWER_OUTPUT_KEYS),
}
ANSWER_PAYLOAD_OUTCOME_ACCEPTED: Final[str] = "accepted"
ANSWER_PAYLOAD_OUTCOME_ACCEPTED_WITH_CITATION_IDS_REMOVED: Final[str] = (
    "accepted_with_citation_ids_removed"
)
ANSWER_PAYLOAD_OUTCOME_REJECTED: Final[str] = "rejected"
FALLBACK_REVIEWER_NOTE: Final[str] = "Model response requires manual review."
FAILURE_REASON_NO_RETRIEVAL: Final[str] = "no_retrieval"
FAILURE_REASON_MODEL_OUTPUT_INVALID: Final[str] = "model_output_invalid"
RETRYABLE_VALIDATION_FAILURE_REASONS: Final[frozenset[str]] = frozenset(
    {
        "payload_shape_invalid",
        "answer_invalid",
        "answer_type_invalid",
        "citation_ids_invalid",
    }
)
FAIL_CLOSED_REVIEWER_NOTE_BY_REASON: Final[Mapping[str, str]] = {
    FAILURE_REASON_NO_RETRIEVAL: "No relevant evidence was retrieved; review manually.",
    FAILURE_REASON_MODEL_OUTPUT_INVALID: (
        "Model output was unusable after retry; review manually."
    ),
    "payload_shape_invalid": "Model output failed validation after retry; review manually.",
    "answer_invalid": "Model answer text was unusable after retry; review manually.",
    "answer_type_invalid": "Model answer type was unusable after retry; review manually.",
    "citation_ids_invalid": "Model citations were unusable after retry; review manually.",
    "no_valid_citations": "All cited evidence was invalid; review manually.",
}

INDEX_ACTION_CREATED: Final[str] = "created"
INDEX_ACTION_REUSED: Final[str] = "reused"
INDEX_ACTION_REBUILT_CONTENT_CHANGE: Final[str] = "rebuilt_content_change"
INDEX_ACTION_REBUILT_INTEGRITY: Final[str] = "rebuilt_integrity"
INDEX_ACTION_BLOCKED: Final[str] = "blocked"
AUTO_REBUILD_INTEGRITY_REASONS: Final[frozenset[str]] = frozenset(
    {
        "collection_empty",
        "chunk_count_missing",
        "chunk_count_mismatch",
        "collection_payload_mismatch",
        "duplicate_logical_chunk_ids",
        "expected_chunk_count_mismatch",
        "logical_chunk_id_mismatch",
        "metadata_chunk_id_missing",
        "metadata_missing",
        "source_coverage_mismatch",
    }
)
BLOCKED_INDEX_REASONS: Final[frozenset[str]] = frozenset(
    {
        "expected_chunks_unavailable",
        "manifest_unavailable",
    }
)

REQUIRED_ENV_VARS: Final[tuple[str, ...]] = ("OPENAI_API_KEY",)

EXPECTED_EVIDENCE_FILE_NAMES: Final[tuple[str, ...]] = (
    SOC2_SUMMARY_FILE_NAME,
    ENCRYPTION_POLICY_FILE_NAME,
    ACCESS_CONTROL_POLICY_FILE_NAME,
    INCIDENT_RESPONSE_POLICY_FILE_NAME,
    BACKUP_RECOVERY_POLICY_FILE_NAME,
)
EXPECTED_QUESTION_IDS: Final[tuple[str, ...]] = tuple(
    f"Q{question_number:02d}" for question_number in range(1, 23)
)
ANSWER_OPENING_TOKENS: Final[tuple[str, ...]] = (
    "Yes.",
    "No.",
    "Partially.",
    "Not stated.",
)
DOCUMENT_TYPE_MARKDOWN: Final[str] = "markdown"
DOCUMENT_TYPE_TEXT: Final[str] = "text"
DOCUMENT_TYPE_PDF: Final[str] = "pdf"
DOCUMENT_TYPE_POLICY: Final[str] = "policy"
TEXT_EVIDENCE_SUFFIXES: Final[tuple[str, ...]] = (".md", ".txt")
CURATED_TEXT_EVIDENCE_FILE_NAMES: Final[tuple[str, ...]] = tuple(
    evidence_file_name
    for evidence_file_name in EXPECTED_EVIDENCE_FILE_NAMES
    if Path(evidence_file_name).suffix.lower() in TEXT_EVIDENCE_SUFFIXES
)
CURATED_PDF_EVIDENCE_FILE_NAMES: Final[tuple[str, ...]] = tuple(
    evidence_file_name
    for evidence_file_name in EXPECTED_EVIDENCE_FILE_NAMES
    if Path(evidence_file_name).suffix.lower() == ".pdf"
)
CHUNK_ID_PREFIX_BY_SOURCE_FILE_NAME: Final[Mapping[str, str]] = {
    ENCRYPTION_POLICY_FILE_NAME: "enc",
    ACCESS_CONTROL_POLICY_FILE_NAME: "acc",
    INCIDENT_RESPONSE_POLICY_FILE_NAME: "ir",
    BACKUP_RECOVERY_POLICY_FILE_NAME: "bkp",
    SOC2_SUMMARY_FILE_NAME: "soc2",
}
CHUNK_METADATA_DOC_TYPE_BY_SOURCE_FILE_NAME: Final[Mapping[str, str]] = {
    ENCRYPTION_POLICY_FILE_NAME: DOCUMENT_TYPE_POLICY,
    ACCESS_CONTROL_POLICY_FILE_NAME: DOCUMENT_TYPE_POLICY,
    INCIDENT_RESPONSE_POLICY_FILE_NAME: DOCUMENT_TYPE_POLICY,
    BACKUP_RECOVERY_POLICY_FILE_NAME: DOCUMENT_TYPE_POLICY,
    SOC2_SUMMARY_FILE_NAME: DOCUMENT_TYPE_PDF,
}
DISPLAY_LABEL_BY_SOURCE_FILE_NAME: Final[Mapping[str, str]] = {
    ENCRYPTION_POLICY_FILE_NAME: "Encryption Policy",
    ACCESS_CONTROL_POLICY_FILE_NAME: "Access Control Policy",
    INCIDENT_RESPONSE_POLICY_FILE_NAME: "Incident Response Policy",
    BACKUP_RECOVERY_POLICY_FILE_NAME: "Backup and Recovery Policy",
    SOC2_SUMMARY_FILE_NAME: "SOC 2 Summary",
}

RUNTIME_DIRECTORIES: Final[tuple[Path, ...]] = (
    RUNTIME_QUESTIONNAIRES_DIR,
    RUNTIME_EVIDENCE_DIR,
    OUTPUTS_DIR,
    CHROMA_DIR,
)
WORKSPACE_HASH_DIRECTORIES: Final[tuple[Path, ...]] = (
    RUNTIME_QUESTIONNAIRES_DIR,
    RUNTIME_EVIDENCE_DIR,
)

SEED_TO_RUNTIME_PATHS: Final[tuple[tuple[Path, Path], ...]] = (
    (
        SEED_QUESTIONNAIRE_DIR / QUESTIONNAIRE_FILE_NAME,
        RUNTIME_QUESTIONNAIRES_DIR / QUESTIONNAIRE_FILE_NAME,
    ),
    *tuple(
        (
            SEED_EVIDENCE_DIR / evidence_file_name,
            RUNTIME_EVIDENCE_DIR / evidence_file_name,
        )
        for evidence_file_name in EXPECTED_EVIDENCE_FILE_NAMES
    ),
)

TESTS_DIR: Final[Path] = REPO_ROOT / "tests"
UNIT_TESTS_DIR: Final[Path] = TESTS_DIR / "unit"
UI_TESTS_DIR: Final[Path] = TESTS_DIR / "ui"
E2E_TESTS_DIR: Final[Path] = TESTS_DIR / "e2e"

TEST_FIXTURES_DIR: Final[Path] = TESTS_DIR / "fixtures"
EXPECTED_OUTCOMES_FIXTURE_PATH: Final[Path] = (
    TEST_FIXTURES_DIR / "expected_outcomes.json"
)
STUBBED_OPENAI_FIXTURES_DIR: Final[Path] = TEST_FIXTURES_DIR / "stubbed_openai"
WORKSPACE_FIXTURES_DIR: Final[Path] = TEST_FIXTURES_DIR / "workspaces"
INDEX_FIXTURES_DIR: Final[Path] = TEST_FIXTURES_DIR / "indexes"

VERIFICATION_ARTIFACTS_DIR: Final[Path] = OUTPUTS_DIR / "verification"
UNIT_TEST_LOGS_DIR: Final[Path] = VERIFICATION_ARTIFACTS_DIR / "unit"
UI_TEST_LOGS_DIR: Final[Path] = VERIFICATION_ARTIFACTS_DIR / "ui"
E2E_TEST_LOGS_DIR: Final[Path] = VERIFICATION_ARTIFACTS_DIR / "e2e"
LIVE_SMOKE_LOGS_DIR: Final[Path] = VERIFICATION_ARTIFACTS_DIR / "live_smoke"

LOG_FORMAT_NAME: Final[str] = "jsonl"
LOG_FILE_EXTENSION: Final[str] = ".jsonl"

LOG_COMPONENT_SETUP: Final[str] = "setup"
LOG_COMPONENT_INDEXING: Final[str] = "indexing"
LOG_COMPONENT_PIPELINE: Final[str] = "pipeline"
LOG_COMPONENT_UI: Final[str] = "ui"
LOG_COMPONENT_EXPORT: Final[str] = "export"
LOG_COMPONENT_VERIFICATION: Final[str] = "verification"
LOG_COMPONENTS: Final[tuple[str, ...]] = (
    LOG_COMPONENT_SETUP,
    LOG_COMPONENT_INDEXING,
    LOG_COMPONENT_PIPELINE,
    LOG_COMPONENT_UI,
    LOG_COMPONENT_EXPORT,
    LOG_COMPONENT_VERIFICATION,
)

LOG_STATUS_STARTED: Final[str] = "started"
LOG_STATUS_COMPLETED: Final[str] = "completed"
LOG_STATUS_FAILED: Final[str] = "failed"
LOG_STATUS_BLOCKED: Final[str] = "blocked"
LOG_STATUS_SKIPPED: Final[str] = "skipped"
LOG_STATUS_RETRYING: Final[str] = "retrying"
LOG_STATUSES: Final[tuple[str, ...]] = (
    LOG_STATUS_STARTED,
    LOG_STATUS_COMPLETED,
    LOG_STATUS_FAILED,
    LOG_STATUS_BLOCKED,
    LOG_STATUS_SKIPPED,
    LOG_STATUS_RETRYING,
)

LOG_LEVELS: Final[tuple[str, ...]] = ("INFO", "WARNING", "ERROR")
LOG_REQUIRED_FIELDS: Final[tuple[str, ...]] = (
    "ts",
    "level",
    "component",
    "event",
    "run_id",
    "status",
    "message",
)
LOG_OPTIONAL_FIELDS: Final[tuple[str, ...]] = (
    "question_id",
    "workspace_hash",
    "manifest_hash",
    "index_action",
    "retrieved_chunk_count",
    "valid_citation_count",
    "retry_attempt",
    "artifact_path",
    "reason",
)


@dataclass
class RuntimeQuestionnaire:
    """In-memory questionnaire rows plus the visible export-column contract."""

    workbook_path: Path
    visible_columns: tuple[str, ...]
    rows: list[dict[str, object]]

    def question_ids(self) -> tuple[str, ...]:
        """Return the loaded question identifiers in workbook order."""
        return tuple(str(row["question_id"]) for row in self.rows)

    def visible_rows(self) -> list[dict[str, object]]:
        """Return only the workbook/export-facing fields in canonical order."""
        return [
            {column_name: row[column_name] for column_name in self.visible_columns}
            for row in self.rows
        ]

    def to_dataframe(self) -> object:
        """Materialize the questionnaire rows as a dataframe with visible columns first."""
        try:
            import pandas as pd
        except ModuleNotFoundError as exc:
            raise ModuleNotFoundError(
                "pandas is required to build the runtime dataframe. Install the project "
                "requirements with `pip install -r requirements.txt` before calling "
                "`RuntimeQuestionnaire.to_dataframe()`."
            ) from exc
        internal_columns = [
            column_name
            for column_name in self.rows[0].keys()
            if column_name not in self.visible_columns
        ]
        return pd.DataFrame(
            self.rows,
            columns=(*self.visible_columns, *internal_columns),
        )


@dataclass(frozen=True)
class EvidenceDocument:
    """One loaded evidence document before normalization and chunking."""

    source_file_name: str
    source_path: Path
    doc_type: str
    text: str
    page_number: int | None = None


@dataclass(frozen=True)
class EvidenceChunk:
    """One deterministic chunk of normalized evidence text."""

    chunk_id: str | None
    source: str
    source_path: Path
    doc_type: str
    text: str
    chunk_number: int
    start_offset: int
    end_offset: int
    section: str | None = None
    page: int | None = None

    def metadata(self) -> dict[str, str | int | None]:
        """Return Chroma-ready metadata for one finalized chunk."""
        if self.chunk_id is None:
            raise ValueError("chunk_id is required before rendering chunk metadata.")
        return {
            "chunk_id": self.chunk_id,
            "source": self.source,
            "doc_type": self.doc_type,
            "section": self.section,
            "page": self.page,
        }


@dataclass(frozen=True)
class RetrievedEvidenceChunk:
    """One retrieved evidence chunk with prompt-ready metadata."""

    chunk_id: str
    source: str
    source_path: Path
    doc_type: str
    text: str
    rank: int
    distance: float | None = None
    section: str | None = None
    page: int | None = None

    def metadata(self) -> dict[str, str | int | None]:
        """Return one stable metadata payload for prompt and citation stages."""
        return {
            "chunk_id": self.chunk_id,
            "source": self.source,
            "doc_type": self.doc_type,
            "section": self.section,
            "page": self.page,
        }


@dataclass(frozen=True)
class ResolvedEvidenceCitation:
    """One reviewer-facing citation with a friendly label and literal snippet text."""

    chunk_id: str
    display_label: str
    snippet_text: str
    source: str
    source_path: Path
    doc_type: str
    section: str | None = None
    page: int | None = None


@dataclass(frozen=True)
class ChromaCollectionHandle:
    """One stable handle to the demo's persistent Chroma collection."""

    collection_name: str
    persist_directory: Path
    client: Any
    collection: Any


@dataclass(frozen=True)
class ChromaReuseStatus:
    """Report whether one persisted Chroma collection can be safely reused."""

    collection_name: str
    persist_directory: Path
    workspace_hash: str | None
    stored_workspace_hash: str | None
    stored_chunk_count: int | None
    actual_chunk_count: int
    index_action: str
    reusable: bool
    reason: str
    collection_handle: ChromaCollectionHandle | None = None


@dataclass(frozen=True)
class ChromaIndexStatus:
    """Report whether the demo index is ready after create/reuse/rebuild decisions."""

    collection_name: str
    persist_directory: Path
    workspace_hash: str | None
    stored_workspace_hash: str | None
    stored_chunk_count: int | None
    actual_chunk_count: int
    index_action: str
    ready: bool
    reason: str
    collection_handle: ChromaCollectionHandle | None = None


@dataclass(frozen=True)
class AnswerPayloadValidationResult:
    """Report whether one model payload is safe for later pipeline stages."""

    usable: bool
    outcome: str
    reviewer_note: str
    answer: str | None = None
    answer_type: str | None = None
    citation_ids: tuple[str, ...] = ()
    invalid_citation_ids: tuple[str, ...] = ()
    failure_reason: str | None = None


@dataclass(frozen=True)
class AnswerConfidenceResult:
    """One deterministic confidence assessment for a validated or fail-closed row."""

    confidence_score: float
    confidence_band: str
    status: str


@dataclass(frozen=True)
class GeneratedAnswerResult:
    """One completed answer decision for a single question and evidence set."""

    answer: str
    answer_type: str
    citation_ids: tuple[str, ...]
    citations: tuple[ResolvedEvidenceCitation, ...]
    confidence_score: float
    confidence_band: str
    status: str
    reviewer_note: str
    invalid_citation_ids: tuple[str, ...] = ()
    failed_closed: bool = False
    failure_reason: str | None = None
    retry_count: int = 0


@dataclass(frozen=True)
class VerificationCommand:
    """One canonical verification command and the conditions around it."""

    name: str
    argv: tuple[str, ...]
    purpose: str
    artifacts_dir: Path | None = None
    requires_openai_api_key: bool = False
    optional: bool = False

    def shell_command(self) -> str:
        """Return the canonical shell string for humans and README output."""
        return shell_join(self.argv)


@dataclass(frozen=True)
class TestSeam:
    """Describe a planned injection or isolation boundary for deterministic tests."""

    name: str
    planned_surface: str
    test_double_strategy: str
    why_it_exists: str


@dataclass(frozen=True)
class LogFieldSpec:
    """Describe one required or optional structured log field."""

    name: str
    required: bool
    example: str
    purpose: str


CANONICAL_VERIFICATION_COMMANDS: Final[tuple[VerificationCommand, ...]] = (
    VerificationCommand(
        name="quick_unit",
        argv=(
            "python",
            "-m",
            "unittest",
            "discover",
            "-s",
            "tests/unit",
            "-p",
            "*_test.py",
            "-v",
        ),
        purpose="Fast confidence check for pure unit-level contracts and helpers.",
        artifacts_dir=UNIT_TEST_LOGS_DIR,
    ),
    VerificationCommand(
        name="ui_suite",
        argv=(
            "python",
            "-m",
            "unittest",
            "discover",
            "-s",
            "tests/ui",
            "-p",
            "*_test.py",
            "-v",
        ),
        purpose="Page-level Streamlit checks for the golden-path operator workflow.",
        artifacts_dir=UI_TEST_LOGS_DIR,
    ),
    VerificationCommand(
        name="deterministic_e2e",
        argv=(
            "python",
            "tests/e2e/run_deterministic_demo.py",
            "--log-dir",
            str(E2E_TEST_LOGS_DIR),
            "--verbose",
        ),
        purpose=(
            "Full local golden-path validation with stubbed model responses and "
            "deterministic evidence expectations."
        ),
        artifacts_dir=E2E_TEST_LOGS_DIR,
    ),
    VerificationCommand(
        name="failure_e2e",
        argv=(
            "python",
            "tests/e2e/run_failure_paths.py",
            "--log-dir",
            str(E2E_TEST_LOGS_DIR),
            "--verbose",
        ),
        purpose=(
            "Failure, retry, review-routing, blocked-state, and index-reuse/rebuild "
            "coverage with deterministic fixtures."
        ),
        artifacts_dir=E2E_TEST_LOGS_DIR,
    ),
    VerificationCommand(
        name="live_smoke",
        argv=(
            "python",
            "tests/e2e/run_live_smoke.py",
            "--question-ids",
            "Q01",
            "Q17",
            "Q21",
            "--log-dir",
            str(LIVE_SMOKE_LOGS_DIR),
            "--verbose",
        ),
        purpose=(
            "Optional real-provider smoke pass for one supported, one partial, and "
            "one unsupported question when OPENAI_API_KEY is available."
        ),
        artifacts_dir=LIVE_SMOKE_LOGS_DIR,
        requires_openai_api_key=True,
        optional=True,
    ),
    VerificationCommand(
        name="closeout_audit",
        argv=("br-closeout-audit", "--issue", "<issue-id>"),
        purpose=(
            "Required audit before closing high-risk implementation beads and "
            "required again after verification-heavy test/logging changes."
        ),
        optional=False,
    ),
)

QUICK_CONFIDENCE_COMMAND_NAMES: Final[tuple[str, ...]] = ("quick_unit",)
FULL_LOCAL_VALIDATION_COMMAND_NAMES: Final[tuple[str, ...]] = (
    "quick_unit",
    "ui_suite",
    "deterministic_e2e",
    "failure_e2e",
)
OPTIONAL_LIVE_VALIDATION_COMMAND_NAMES: Final[tuple[str, ...]] = ("live_smoke",)

CLOSEOUT_AUDIT_RULES: Final[tuple[str, ...]] = (
    "Run br-closeout-audit before closing any high-risk bead that changes runtime, retrieval, scoring, or export behavior.",
    "Run br-closeout-audit again after closing verification-heavy work that adds or materially changes tests, e2e scripts, or diagnostic logging.",
)

TEST_SEAMS: Final[tuple[TestSeam, ...]] = (
    TestSeam(
        name="workspace_setup",
        planned_surface=(
            "generate_demo_data.py functions that create runtime directories, copy "
            "seed files, compute manifests, and reset outputs from explicit root paths"
        ),
        test_double_strategy=(
            "Use temporary directories and fixture trees rather than mutating the "
            "real repo workspace."
        ),
        why_it_exists=(
            "Workspace setup is filesystem-heavy and must be repeatable, idempotent, "
            "and safe to exercise in unit and e2e tests."
        ),
    ),
    TestSeam(
        name="openai_embeddings",
        planned_surface=(
            "rag.py adapter or factory that produces embeddings independently from "
            "answer generation"
        ),
        test_double_strategy=(
            "Inject deterministic fake embedding responses or fixed vectors without "
            "calling the network."
        ),
        why_it_exists=(
            "Indexing and retrieval tests must stay deterministic and must not depend "
            "on a live provider."
        ),
    ),
    TestSeam(
        name="openai_answers",
        planned_surface=(
            "rag.py adapter or factory that receives prompt payloads and returns the "
            "structured answer contract"
        ),
        test_double_strategy=(
            "Swap in scripted payload fixtures for supported, partial, unsupported, "
            "malformed, and retry-triggering responses."
        ),
        why_it_exists=(
            "Prompting, validation, scoring, and fail-closed behavior need exact, "
            "repeatable answer payloads."
        ),
    ),
    TestSeam(
        name="chroma_index",
        planned_surface=(
            "rag.py client or collection boundary that isolates persistence, reuse, "
            "rebuild, and integrity checks from chunk construction"
        ),
        test_double_strategy=(
            "Use temp persistence roots or fake collection objects to force reuse, "
            "stale-state, and rebuild scenarios deterministically."
        ),
        why_it_exists=(
            "The index lifecycle is central to correctness but should be testable "
            "without relying on implicit ambient state."
        ),
    ),
    TestSeam(
        name="result_shaping",
        planned_surface=(
            "Pure helpers in rag.py that assemble canonical result rows, evidence "
            "labels, reviewer notes, export values, and review ordering"
        ),
        test_double_strategy=(
            "Call helpers directly with small fixtures and assert exact field values, "
            "ordering, and fail-closed defaults."
        ),
        why_it_exists=(
            "Reviewer-facing outputs must stay stable across UI, exports, and tests."
        ),
    ),
    TestSeam(
        name="streamlit_orchestration",
        planned_surface=(
            "app.py event handlers that translate session state into explicit calls "
            "for setup, run, selection, review, and export actions"
        ),
        test_double_strategy=(
            "Drive UI tests through Streamlit's testing harness with fake workspace "
            "state and deterministic pipeline responses."
        ),
        why_it_exists=(
            "The UI should be testable without relying on manual clicks or a live "
            "provider session."
        ),
    ),
)

LOG_FIELD_SPECS: Final[tuple[LogFieldSpec, ...]] = (
    LogFieldSpec(
        name="ts",
        required=True,
        example="2026-04-27T19:00:00Z",
        purpose="UTC timestamp for stable event ordering across scripts and runs.",
    ),
    LogFieldSpec(
        name="level",
        required=True,
        example="INFO",
        purpose="Severity for filtering happy-path, warning, and failure events.",
    ),
    LogFieldSpec(
        name="component",
        required=True,
        example=LOG_COMPONENT_PIPELINE,
        purpose="High-level subsystem responsible for the event.",
    ),
    LogFieldSpec(
        name="event",
        required=True,
        example="row_completed",
        purpose="Stable machine-friendly event name for scripts and audits.",
    ),
    LogFieldSpec(
        name="run_id",
        required=True,
        example="demo-run-001",
        purpose="Correlation key spanning one setup or answer-generation run.",
    ),
    LogFieldSpec(
        name="status",
        required=True,
        example=LOG_STATUS_COMPLETED,
        purpose="Outcome marker for start, completion, failure, block, skip, or retry.",
    ),
    LogFieldSpec(
        name="message",
        required=True,
        example="Answered Q01 with 2 valid citations.",
        purpose="Human-readable summary of the event in one line.",
    ),
    LogFieldSpec(
        name="question_id",
        required=False,
        example="Q01",
        purpose="Question-level correlation for retrieval, scoring, and review events.",
    ),
    LogFieldSpec(
        name="workspace_hash",
        required=False,
        example="abc123workspacehash",
        purpose="Workspace identity for reuse, rebuild, and setup decisions.",
    ),
    LogFieldSpec(
        name="manifest_hash",
        required=False,
        example="def456manifesthash",
        purpose="Manifest identity when comparing persisted state to current files.",
    ),
    LogFieldSpec(
        name="index_action",
        required=False,
        example=INDEX_ACTION_REUSED,
        purpose="Explicit created/reused/rebuilt/blocked index decision outcome.",
    ),
    LogFieldSpec(
        name="retrieved_chunk_count",
        required=False,
        example="5",
        purpose="Retrieved evidence volume before answer validation.",
    ),
    LogFieldSpec(
        name="valid_citation_count",
        required=False,
        example="2",
        purpose="Valid citation count after payload validation and dedupe.",
    ),
    LogFieldSpec(
        name="retry_attempt",
        required=False,
        example="1",
        purpose="Retry counter for malformed or schema-invalid model responses.",
    ),
    LogFieldSpec(
        name="artifact_path",
        required=False,
        example="data/outputs/Answered_Questionnaire.xlsx",
        purpose="Published artifact path for exports and saved diagnostics.",
    ),
    LogFieldSpec(
        name="reason",
        required=False,
        example="missing_openai_api_key",
        purpose="Compact machine-friendly explanation for blocked or fail-closed states.",
    ),
)

LOGGING_CONVENTIONS: Final[tuple[str, ...]] = (
    "Emit one structured JSONL record per significant phase or question-level event.",
    "Always include the required fields and only include optional fields when they add real diagnostic value.",
    "Use question_id for row-level events, workspace_hash or manifest_hash for reuse decisions, and artifact_path for published outputs.",
    "Do not log API keys, raw secrets, or full provider transcripts by default.",
)

CONSERVATIVE_ANSWER_SYSTEM_PROMPT: Final[str] = "\n".join(
    (
        "You answer one security questionnaire row using only the provided evidence chunks.",
        "Answer only from the provided evidence chunks.",
        "Do not use outside knowledge.",
        "Do not invent policies, timelines, capabilities, or citations.",
        "Return valid JSON with exactly these keys: answer, answer_type, citation_ids, reviewer_note.",
        "answer_type must be one of: supported, partial, unsupported.",
        "If evidence is incomplete, use partial.",
        "If evidence is missing, use unsupported.",
        "The answer must be 1 to 3 sentences.",
        "The answer must begin with exactly one of: Yes., No., Partially., Not stated.",
        "Keep the answer short and readable.",
        "Use only chunk_id values from the provided evidence list in citation_ids.",
        "Do not include raw filenames in the answer unless necessary.",
        "Keep reviewer_note short and reviewer-oriented.",
    )
)


def question_order_index(question_id: str) -> int:
    """Return the canonical questionnaire order for a known question identifier."""
    return EXPECTED_QUESTION_IDS.index(question_id)


def confidence_band_for_score(confidence_score: float) -> str:
    """Map a numeric confidence score into the planned display band."""
    if confidence_score >= HIGH_CONFIDENCE_THRESHOLD:
        return CONFIDENCE_BAND_HIGH
    if confidence_score >= MEDIUM_CONFIDENCE_THRESHOLD:
        return CONFIDENCE_BAND_MEDIUM
    return CONFIDENCE_BAND_LOW


def review_status_for_score(confidence_score: float) -> str:
    """Map a numeric confidence score into the planned review status."""
    if confidence_score >= REVIEW_QUEUE_THRESHOLD:
        return STATUS_READY_FOR_REVIEW
    return STATUS_NEEDS_REVIEW


def score_answer_confidence(
    answer_type: str,
    *,
    valid_citation_count: int,
) -> AnswerConfidenceResult:
    """Score one validated answer type and citation count using the planned rules."""
    if answer_type not in ANSWER_TYPES:
        raise ValueError(f"Unsupported answer_type {answer_type!r} for scoring.")
    if valid_citation_count < 0:
        raise ValueError("valid_citation_count must be zero or greater.")

    if valid_citation_count == 0:
        confidence_score = FAIL_CLOSED_SCORE
    elif answer_type == ANSWER_TYPE_SUPPORTED:
        confidence_score = (
            SUPPORTED_WITH_TWO_PLUS_CITATIONS_SCORE
            if valid_citation_count >= 2
            else SUPPORTED_WITH_ONE_CITATION_SCORE
        )
    elif answer_type == ANSWER_TYPE_PARTIAL:
        confidence_score = PARTIAL_SCORE
    else:
        confidence_score = UNSUPPORTED_SCORE

    return AnswerConfidenceResult(
        confidence_score=confidence_score,
        confidence_band=confidence_band_for_score(confidence_score),
        status=review_status_for_score(confidence_score),
    )


def build_evidence_display_value(display_labels: Sequence[str]) -> str:
    """Render reviewer-facing evidence labels for workbook export."""
    return "; ".join(label for label in display_labels if label)


def _friendly_source_label(source_file_name: str) -> str:
    """Return the stable human-facing label for one evidence source."""
    return DISPLAY_LABEL_BY_SOURCE_FILE_NAME.get(
        source_file_name,
        source_file_name.replace("_", " ").rsplit(".", 1)[0],
    )


def build_citation_display_label(
    source_file_name: str,
    section: str | None = None,
    page: int | None = None,
) -> str:
    """Create the reviewer-facing label for one citation."""
    base_label = _friendly_source_label(source_file_name)
    if section and section != base_label:
        return f"{base_label} — {section}"
    if page is not None:
        return f"{base_label} — Page {page}"
    return base_label


def make_result_row_defaults() -> dict[str, object]:
    """Return the canonical default shape for internal result-row fields."""
    return {
        "answer": "",
        "answer_type": "",
        "citation_ids": [],
        "citations": [],
        "confidence_score": 0.0,
        "confidence_band": "",
        "status": "",
        "reviewer_note": "",
        "evidence_labels": [],
        "index_action": "",
        "run_id": "",
    }


def runtime_questionnaire_path() -> Path:
    """Return the curated runtime questionnaire path inside the workspace."""
    return RUNTIME_QUESTIONNAIRES_DIR / QUESTIONNAIRE_FILE_NAME


def runtime_evidence_directory() -> Path:
    """Return the curated runtime evidence directory inside the workspace."""
    return RUNTIME_EVIDENCE_DIR


def runtime_manifest_path() -> Path:
    """Return the current runtime workspace-manifest path."""
    return DATA_DIR / MANIFEST_FILE_NAME


def answered_questionnaire_output_path(output_dir: Path | None = None) -> Path:
    """Return the canonical workbook export path for one completed run."""
    base_output_dir = OUTPUTS_DIR if output_dir is None else Path(output_dir)
    return base_output_dir / ANSWERED_QUESTIONNAIRE_FILE_NAME


def current_workspace_hash(manifest_path: Path | None = None) -> str:
    """Read the current runtime workspace hash from the manifest file."""
    path = manifest_path or runtime_manifest_path()
    if not path.exists():
        raise FileNotFoundError(
            "Runtime workspace manifest is missing at "
            f"{path}. Run `python generate_demo_data.py` before indexing."
        )

    payload = json.loads(path.read_text(encoding="utf-8"))
    workspace_hash = payload.get("workspace_hash")
    if not isinstance(workspace_hash, str) or not workspace_hash.strip():
        raise ValueError(
            f"Runtime workspace manifest {path} is missing a valid `workspace_hash`."
        )
    return workspace_hash


def chroma_persist_directory(persist_directory: Path | None = None) -> Path:
    """Return the stable local persistence directory for the Chroma demo index."""
    if persist_directory is None:
        return CHROMA_DIR.resolve()
    return Path(persist_directory).expanduser().resolve()


def _import_chromadb() -> Any:
    """Import chromadb lazily so non-indexing shells can still load rag helpers."""
    try:
        import chromadb
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            "chromadb is required for the persistent demo index. "
            "Install dependencies with `pip install -r requirements.txt`."
        ) from exc
    return chromadb


def _collection_metadata(handle: ChromaCollectionHandle) -> dict[str, Any]:
    """Return one mutable metadata mapping for the collection handle."""
    metadata = handle.collection.metadata or {}
    if not isinstance(metadata, dict):
        return {}
    return dict(metadata)


def _existing_collection_names(client: Any) -> set[str]:
    """Return the existing collection names for one Chroma client."""
    return {
        str(getattr(collection_item, "name", collection_item))
        for collection_item in client.list_collections()
    }


def get_existing_chroma_collection(
    *,
    collection_name: str = COLLECTION_NAME,
    persist_directory: Path | None = None,
) -> ChromaCollectionHandle | None:
    """Open one existing persistent Chroma collection without creating a new one."""
    normalized_collection_name = collection_name.strip()
    if not normalized_collection_name:
        raise ValueError("collection_name must be a non-empty string.")

    persistence_path = chroma_persist_directory(persist_directory)
    if persistence_path.exists() and not persistence_path.is_dir():
        raise ValueError("persist_directory must resolve to a directory path.")
    if not persistence_path.exists():
        return None

    chromadb = _import_chromadb()
    client = chromadb.PersistentClient(path=str(persistence_path))
    if normalized_collection_name not in _existing_collection_names(client):
        return None

    collection = client.get_collection(name=normalized_collection_name)
    return ChromaCollectionHandle(
        collection_name=normalized_collection_name,
        persist_directory=persistence_path,
        client=client,
        collection=collection,
    )


def get_or_create_chroma_collection(
    *,
    collection_name: str = COLLECTION_NAME,
    persist_directory: Path | None = None,
) -> ChromaCollectionHandle:
    """Create or reconnect to one persistent Chroma collection."""
    normalized_collection_name = collection_name.strip()
    if not normalized_collection_name:
        raise ValueError("collection_name must be a non-empty string.")

    persistence_path = chroma_persist_directory(persist_directory)
    if persistence_path.exists() and not persistence_path.is_dir():
        raise ValueError("persist_directory must resolve to a directory path.")
    persistence_path.mkdir(parents=True, exist_ok=True)

    chromadb = _import_chromadb()
    client = chromadb.PersistentClient(path=str(persistence_path))
    collection = client.get_or_create_collection(name=normalized_collection_name)
    return ChromaCollectionHandle(
        collection_name=normalized_collection_name,
        persist_directory=persistence_path,
        client=client,
        collection=collection,
    )


def get_or_create_demo_chroma_collection(
    persist_directory: Path | None = None,
) -> ChromaCollectionHandle:
    """Create or reconnect to the one canonical demo collection."""
    return get_or_create_chroma_collection(
        collection_name=COLLECTION_NAME,
        persist_directory=persist_directory,
    )


def record_indexed_workspace_state(
    collection_handle: ChromaCollectionHandle,
    *,
    workspace_hash: str,
    chunk_count: int,
) -> dict[str, Any]:
    """Record the last indexed workspace identity on the Chroma collection."""
    metadata = _collection_metadata(collection_handle)
    metadata["workspace_hash"] = workspace_hash
    metadata["chunk_count"] = chunk_count
    collection_handle.collection.modify(metadata=metadata)
    return metadata


def _expected_curated_chunk_inventory(
    evidence_dir: Path | None = None,
) -> tuple[int, frozenset[str], frozenset[str]]:
    """Return the deterministic chunk-count, chunk-id, and source inventory."""
    expected_chunks = build_curated_evidence_chunks(evidence_dir)
    expected_chunk_ids = tuple(chunk.chunk_id for chunk in expected_chunks)
    if any(chunk_id is None for chunk_id in expected_chunk_ids):
        raise ValueError("Curated evidence chunks must provide chunk IDs for integrity checks.")

    return (
        len(expected_chunks),
        frozenset(str(chunk_id) for chunk_id in expected_chunk_ids),
        frozenset(chunk.source for chunk in expected_chunks),
    )


def _stored_collection_payload(
    collection_handle: ChromaCollectionHandle,
) -> tuple[tuple[str, ...], tuple[dict[str, Any] | None, ...]]:
    """Return one storage payload snapshot for integrity inspection."""
    payload = collection_handle.collection.get(include=["metadatas"])
    raw_ids = payload.get("ids")
    raw_metadatas = payload.get("metadatas")
    stored_ids = tuple(str(raw_id) for raw_id in raw_ids) if isinstance(raw_ids, list) else ()
    stored_metadatas = (
        tuple(raw_metadata if isinstance(raw_metadata, dict) else None for raw_metadata in raw_metadatas)
        if isinstance(raw_metadatas, list)
        else ()
    )
    return stored_ids, stored_metadatas


def evaluate_chroma_reuse(
    *,
    collection_name: str = COLLECTION_NAME,
    persist_directory: Path | None = None,
    evidence_dir: Path | None = None,
    manifest_path: Path | None = None,
) -> ChromaReuseStatus:
    """Return whether the existing collection can be safely reused."""
    try:
        workspace_hash = current_workspace_hash(manifest_path)
    except (FileNotFoundError, ValueError):
        return ChromaReuseStatus(
            collection_name=collection_name,
            persist_directory=chroma_persist_directory(persist_directory),
            workspace_hash=None,
            stored_workspace_hash=None,
            stored_chunk_count=None,
            actual_chunk_count=0,
            index_action=INDEX_ACTION_BLOCKED,
            reusable=False,
            reason="manifest_unavailable",
            collection_handle=None,
        )

    collection_handle = get_existing_chroma_collection(
        collection_name=collection_name,
        persist_directory=persist_directory,
    )
    if collection_handle is None:
        return ChromaReuseStatus(
            collection_name=collection_name,
            persist_directory=chroma_persist_directory(persist_directory),
            workspace_hash=workspace_hash,
            stored_workspace_hash=None,
            stored_chunk_count=None,
            actual_chunk_count=0,
            index_action=INDEX_ACTION_BLOCKED,
            reusable=False,
            reason="collection_missing",
            collection_handle=None,
        )

    metadata = _collection_metadata(collection_handle)
    stored_workspace_hash = metadata.get("workspace_hash")
    raw_stored_chunk_count = metadata.get("chunk_count")
    stored_chunk_count = (
        raw_stored_chunk_count
        if isinstance(raw_stored_chunk_count, int) and raw_stored_chunk_count >= 0
        else None
    )
    try:
        actual_chunk_count = int(collection_handle.collection.count())
    except Exception:
        return ChromaReuseStatus(
            collection_name=collection_name,
            persist_directory=collection_handle.persist_directory,
            workspace_hash=workspace_hash,
            stored_workspace_hash=(
                stored_workspace_hash if isinstance(stored_workspace_hash, str) else None
            ),
            stored_chunk_count=stored_chunk_count,
            actual_chunk_count=0,
            index_action=INDEX_ACTION_BLOCKED,
            reusable=False,
            reason="collection_payload_mismatch",
            collection_handle=collection_handle,
        )

    if stored_workspace_hash != workspace_hash:
        return ChromaReuseStatus(
            collection_name=collection_name,
            persist_directory=collection_handle.persist_directory,
            workspace_hash=workspace_hash,
            stored_workspace_hash=(
                stored_workspace_hash if isinstance(stored_workspace_hash, str) else None
            ),
            stored_chunk_count=stored_chunk_count,
            actual_chunk_count=actual_chunk_count,
            index_action=INDEX_ACTION_BLOCKED,
            reusable=False,
            reason="workspace_hash_mismatch",
            collection_handle=collection_handle,
        )

    if stored_chunk_count is None:
        return ChromaReuseStatus(
            collection_name=collection_name,
            persist_directory=collection_handle.persist_directory,
            workspace_hash=workspace_hash,
            stored_workspace_hash=workspace_hash,
            stored_chunk_count=None,
            actual_chunk_count=actual_chunk_count,
            index_action=INDEX_ACTION_BLOCKED,
            reusable=False,
            reason="chunk_count_missing",
            collection_handle=collection_handle,
        )

    if actual_chunk_count == 0:
        return ChromaReuseStatus(
            collection_name=collection_name,
            persist_directory=collection_handle.persist_directory,
            workspace_hash=workspace_hash,
            stored_workspace_hash=workspace_hash,
            stored_chunk_count=stored_chunk_count,
            actual_chunk_count=actual_chunk_count,
            index_action=INDEX_ACTION_BLOCKED,
            reusable=False,
            reason="collection_empty",
            collection_handle=collection_handle,
        )

    if actual_chunk_count != stored_chunk_count:
        return ChromaReuseStatus(
            collection_name=collection_name,
            persist_directory=collection_handle.persist_directory,
            workspace_hash=workspace_hash,
            stored_workspace_hash=workspace_hash,
            stored_chunk_count=stored_chunk_count,
            actual_chunk_count=actual_chunk_count,
            index_action=INDEX_ACTION_BLOCKED,
            reusable=False,
            reason="chunk_count_mismatch",
            collection_handle=collection_handle,
        )

    try:
        (
            expected_chunk_count,
            expected_chunk_ids,
            expected_sources,
        ) = _expected_curated_chunk_inventory(evidence_dir)
    except (FileNotFoundError, ModuleNotFoundError, ValueError):
        return ChromaReuseStatus(
            collection_name=collection_name,
            persist_directory=collection_handle.persist_directory,
            workspace_hash=workspace_hash,
            stored_workspace_hash=workspace_hash,
            stored_chunk_count=stored_chunk_count,
            actual_chunk_count=actual_chunk_count,
            index_action=INDEX_ACTION_BLOCKED,
            reusable=False,
            reason="expected_chunks_unavailable",
            collection_handle=collection_handle,
        )

    if actual_chunk_count != expected_chunk_count:
        return ChromaReuseStatus(
            collection_name=collection_name,
            persist_directory=collection_handle.persist_directory,
            workspace_hash=workspace_hash,
            stored_workspace_hash=workspace_hash,
            stored_chunk_count=stored_chunk_count,
            actual_chunk_count=actual_chunk_count,
            index_action=INDEX_ACTION_BLOCKED,
            reusable=False,
            reason="expected_chunk_count_mismatch",
            collection_handle=collection_handle,
        )

    try:
        stored_ids, stored_metadatas = _stored_collection_payload(collection_handle)
    except Exception:
        return ChromaReuseStatus(
            collection_name=collection_name,
            persist_directory=collection_handle.persist_directory,
            workspace_hash=workspace_hash,
            stored_workspace_hash=workspace_hash,
            stored_chunk_count=stored_chunk_count,
            actual_chunk_count=actual_chunk_count,
            index_action=INDEX_ACTION_BLOCKED,
            reusable=False,
            reason="collection_payload_mismatch",
            collection_handle=collection_handle,
        )
    if (
        len(stored_ids) != actual_chunk_count
        or len(stored_metadatas) != actual_chunk_count
        or frozenset(stored_ids) != expected_chunk_ids
    ):
        return ChromaReuseStatus(
            collection_name=collection_name,
            persist_directory=collection_handle.persist_directory,
            workspace_hash=workspace_hash,
            stored_workspace_hash=workspace_hash,
            stored_chunk_count=stored_chunk_count,
            actual_chunk_count=actual_chunk_count,
            index_action=INDEX_ACTION_BLOCKED,
            reusable=False,
            reason="collection_payload_mismatch",
            collection_handle=collection_handle,
        )

    if any(metadata is None for metadata in stored_metadatas):
        return ChromaReuseStatus(
            collection_name=collection_name,
            persist_directory=collection_handle.persist_directory,
            workspace_hash=workspace_hash,
            stored_workspace_hash=workspace_hash,
            stored_chunk_count=stored_chunk_count,
            actual_chunk_count=actual_chunk_count,
            index_action=INDEX_ACTION_BLOCKED,
            reusable=False,
            reason="metadata_missing",
            collection_handle=collection_handle,
        )

    logical_chunk_ids: list[str] = []
    stored_sources: list[str] = []
    for metadata in stored_metadatas:
        if metadata is None:
            continue

        logical_chunk_id = metadata.get("chunk_id")
        if not isinstance(logical_chunk_id, str) or not logical_chunk_id.strip():
            return ChromaReuseStatus(
                collection_name=collection_name,
                persist_directory=collection_handle.persist_directory,
                workspace_hash=workspace_hash,
                stored_workspace_hash=workspace_hash,
                stored_chunk_count=stored_chunk_count,
                actual_chunk_count=actual_chunk_count,
                index_action=INDEX_ACTION_BLOCKED,
                reusable=False,
                reason="metadata_chunk_id_missing",
                collection_handle=collection_handle,
            )
        logical_chunk_ids.append(logical_chunk_id)

        source_name = metadata.get("source")
        if not isinstance(source_name, str) or not source_name.strip():
            return ChromaReuseStatus(
                collection_name=collection_name,
                persist_directory=collection_handle.persist_directory,
                workspace_hash=workspace_hash,
                stored_workspace_hash=workspace_hash,
                stored_chunk_count=stored_chunk_count,
                actual_chunk_count=actual_chunk_count,
                index_action=INDEX_ACTION_BLOCKED,
                reusable=False,
                reason="source_coverage_mismatch",
                collection_handle=collection_handle,
            )
        stored_sources.append(source_name)

    if len(set(logical_chunk_ids)) != len(logical_chunk_ids):
        return ChromaReuseStatus(
            collection_name=collection_name,
            persist_directory=collection_handle.persist_directory,
            workspace_hash=workspace_hash,
            stored_workspace_hash=workspace_hash,
            stored_chunk_count=stored_chunk_count,
            actual_chunk_count=actual_chunk_count,
            index_action=INDEX_ACTION_BLOCKED,
            reusable=False,
            reason="duplicate_logical_chunk_ids",
            collection_handle=collection_handle,
        )

    if frozenset(logical_chunk_ids) != expected_chunk_ids:
        return ChromaReuseStatus(
            collection_name=collection_name,
            persist_directory=collection_handle.persist_directory,
            workspace_hash=workspace_hash,
            stored_workspace_hash=workspace_hash,
            stored_chunk_count=stored_chunk_count,
            actual_chunk_count=actual_chunk_count,
            index_action=INDEX_ACTION_BLOCKED,
            reusable=False,
            reason="logical_chunk_id_mismatch",
            collection_handle=collection_handle,
        )

    if frozenset(stored_sources) != expected_sources:
        return ChromaReuseStatus(
            collection_name=collection_name,
            persist_directory=collection_handle.persist_directory,
            workspace_hash=workspace_hash,
            stored_workspace_hash=workspace_hash,
            stored_chunk_count=stored_chunk_count,
            actual_chunk_count=actual_chunk_count,
            index_action=INDEX_ACTION_BLOCKED,
            reusable=False,
            reason="source_coverage_mismatch",
            collection_handle=collection_handle,
        )

    return ChromaReuseStatus(
        collection_name=collection_name,
        persist_directory=collection_handle.persist_directory,
        workspace_hash=workspace_hash,
        stored_workspace_hash=workspace_hash,
        stored_chunk_count=stored_chunk_count,
        actual_chunk_count=actual_chunk_count,
        index_action=INDEX_ACTION_REUSED,
        reusable=True,
        reason="reused",
        collection_handle=collection_handle,
    )


def _index_status_from_reuse_status(
    reuse_status: ChromaReuseStatus,
) -> ChromaIndexStatus:
    """Adapt one reuse-only status into the generic index decision surface."""
    return ChromaIndexStatus(
        collection_name=reuse_status.collection_name,
        persist_directory=reuse_status.persist_directory,
        workspace_hash=reuse_status.workspace_hash,
        stored_workspace_hash=reuse_status.stored_workspace_hash,
        stored_chunk_count=reuse_status.stored_chunk_count,
        actual_chunk_count=reuse_status.actual_chunk_count,
        index_action=reuse_status.index_action,
        ready=reuse_status.reusable,
        reason=reuse_status.reason,
        collection_handle=reuse_status.collection_handle,
    )


def _normalize_workbook_text_cell(value: object) -> str:
    """Convert a workbook cell value into the canonical in-memory text form."""
    if value is None:
        return ""
    return str(value).strip()


def load_text_evidence_document(source_path: Path) -> EvidenceDocument:
    """Load one markdown or plain-text evidence file into the common document shape."""
    suffix = source_path.suffix.lower()
    if suffix not in TEXT_EVIDENCE_SUFFIXES:
        supported_suffixes = ", ".join(TEXT_EVIDENCE_SUFFIXES)
        raise ValueError(
            f"Unsupported text evidence file type for {source_path.name}: "
            f"expected one of {supported_suffixes}."
        )

    doc_type = DOCUMENT_TYPE_MARKDOWN if suffix == ".md" else DOCUMENT_TYPE_TEXT
    return EvidenceDocument(
        source_file_name=source_path.name,
        source_path=source_path,
        doc_type=doc_type,
        text=source_path.read_text(encoding="utf-8"),
    )


def load_curated_text_evidence_documents(
    evidence_dir: Path | None = None,
) -> tuple[EvidenceDocument, ...]:
    """Load the four curated markdown policy files in stable runtime order."""
    base_directory = evidence_dir or runtime_evidence_directory()
    return tuple(
        load_text_evidence_document(base_directory / evidence_file_name)
        for evidence_file_name in CURATED_TEXT_EVIDENCE_FILE_NAMES
    )


def load_pdf_evidence_pages(source_path: Path) -> tuple[EvidenceDocument, ...]:
    """Load one text-based PDF into page-aware evidence documents via pypdf."""
    if source_path.suffix.lower() != ".pdf":
        raise ValueError(
            f"Unsupported PDF evidence file type for {source_path.name}: expected .pdf."
        )

    try:
        from pypdf import PdfReader
    except ModuleNotFoundError as exc:
        raise ModuleNotFoundError(
            "pypdf is required to extract text from the bundled SOC 2 PDF. Install the "
            "project requirements with `pip install -r requirements.txt` before calling "
            "`load_pdf_evidence_pages()`."
        ) from exc

    reader = PdfReader(str(source_path))
    page_documents: list[EvidenceDocument] = []
    for page_number, page in enumerate(reader.pages, start=1):
        page_text = page.extract_text() or ""
        if not page_text.strip():
            continue
        page_documents.append(
            EvidenceDocument(
                source_file_name=source_path.name,
                source_path=source_path,
                doc_type=DOCUMENT_TYPE_PDF,
                text=page_text,
                page_number=page_number,
            )
        )

    if not page_documents:
        raise ValueError(
            f"PDF evidence file {source_path.name} did not yield any extractable text "
            "via pypdf."
        )
    return tuple(page_documents)


def load_curated_pdf_evidence_documents(
    evidence_dir: Path | None = None,
) -> tuple[EvidenceDocument, ...]:
    """Load the curated SOC 2 PDF into page-aware evidence documents."""
    base_directory = evidence_dir or runtime_evidence_directory()
    page_documents: list[EvidenceDocument] = []
    for evidence_file_name in CURATED_PDF_EVIDENCE_FILE_NAMES:
        page_documents.extend(load_pdf_evidence_pages(base_directory / evidence_file_name))
    return tuple(page_documents)


def load_curated_evidence_documents(
    evidence_dir: Path | None = None,
) -> tuple[EvidenceDocument, ...]:
    """Load the full curated evidence pack into one stable document sequence."""
    return (
        *load_curated_text_evidence_documents(evidence_dir),
        *load_curated_pdf_evidence_documents(evidence_dir),
    )


def normalize_evidence_text(text: str) -> str:
    """Normalize text deterministically while preserving headings and structure."""
    normalized_line_endings = text.replace("\r\n", "\n").replace("\r", "\n")
    normalized_lines: list[str] = []
    previous_line_blank = False
    for raw_line in normalized_line_endings.split("\n"):
        line = raw_line.rstrip()
        line_is_blank = line == ""
        if line_is_blank and previous_line_blank:
            continue
        normalized_lines.append(line)
        previous_line_blank = line_is_blank
    return "\n".join(normalized_lines).strip("\n")


def normalize_evidence_document(document: EvidenceDocument) -> EvidenceDocument:
    """Return one evidence document with normalized text and unchanged metadata."""
    return EvidenceDocument(
        source_file_name=document.source_file_name,
        source_path=document.source_path,
        doc_type=document.doc_type,
        text=normalize_evidence_text(document.text),
        page_number=document.page_number,
    )


def normalize_evidence_documents(
    documents: Sequence[EvidenceDocument],
) -> tuple[EvidenceDocument, ...]:
    """Normalize a sequence of loaded evidence documents in stable order."""
    return tuple(normalize_evidence_document(document) for document in documents)


def _chunk_boundaries(
    text: str,
    *,
    chunk_size: int = CHUNK_SIZE_CHARS,
    chunk_overlap: int = CHUNK_OVERLAP_CHARS,
) -> tuple[tuple[int, int], ...]:
    """Return deterministic inclusive-exclusive chunk boundaries for one text."""
    if chunk_size <= 0:
        raise ValueError("chunk_size must be a positive integer.")
    if chunk_overlap < 0:
        raise ValueError("chunk_overlap cannot be negative.")
    if chunk_overlap >= chunk_size:
        raise ValueError("chunk_overlap must be smaller than chunk_size.")
    if not text:
        return ()

    chunk_step = chunk_size - chunk_overlap
    boundaries: list[tuple[int, int]] = []
    start_offset = 0
    while start_offset < len(text):
        end_offset = min(start_offset + chunk_size, len(text))
        boundaries.append((start_offset, end_offset))
        if end_offset >= len(text):
            break
        start_offset += chunk_step
    return tuple(boundaries)


def _chunk_id_prefix_for_source(source_file_name: str) -> str:
    """Return the stable chunk-id prefix for one curated evidence source."""
    try:
        return CHUNK_ID_PREFIX_BY_SOURCE_FILE_NAME[source_file_name]
    except KeyError as exc:
        raise ValueError(
            f"Unsupported chunk-id source file {source_file_name!r} for the curated demo."
        ) from exc


def _chunk_metadata_doc_type_for_source(source_file_name: str) -> str:
    """Return the metadata doc_type for one curated evidence source."""
    try:
        return CHUNK_METADATA_DOC_TYPE_BY_SOURCE_FILE_NAME[source_file_name]
    except KeyError as exc:
        raise ValueError(
            f"Unsupported metadata doc_type source file {source_file_name!r} for the "
            "curated demo."
        ) from exc


def _markdown_heading_text(line: str) -> str | None:
    """Return one markdown ATX heading text when the line is a heading."""
    stripped_line = line.strip()
    if not stripped_line.startswith("#"):
        return None
    heading_marks, separator, heading_text = stripped_line.partition(" ")
    if not separator or set(heading_marks) != {"#"}:
        return None
    heading_text = heading_text.strip()
    return heading_text or None


def _section_label_for_chunk(document: EvidenceDocument, start_offset: int) -> str | None:
    """Return the nearest section label or a stable fallback for one chunk start."""
    if document.doc_type == DOCUMENT_TYPE_PDF:
        return None

    nearest_heading: str | None = None
    line_start_offset = 0
    for line in document.text.split("\n"):
        if line_start_offset > start_offset:
            break
        heading_text = _markdown_heading_text(line)
        if heading_text:
            nearest_heading = heading_text
        line_start_offset += len(line) + 1
    return nearest_heading or _friendly_source_label(document.source_file_name)


def chunk_evidence_document(
    document: EvidenceDocument,
    *,
    chunk_size: int = CHUNK_SIZE_CHARS,
    chunk_overlap: int = CHUNK_OVERLAP_CHARS,
) -> tuple[EvidenceChunk, ...]:
    """Split one normalized evidence document into deterministic overlapping chunks."""
    boundaries = _chunk_boundaries(
        document.text,
        chunk_size=chunk_size,
        chunk_overlap=chunk_overlap,
    )
    return tuple(
        EvidenceChunk(
            chunk_id=None,
            source=document.source_file_name,
            source_path=document.source_path,
            doc_type=_chunk_metadata_doc_type_for_source(document.source_file_name),
            text=document.text[start_offset:end_offset],
            chunk_number=chunk_number,
            start_offset=start_offset,
            end_offset=end_offset,
            section=_section_label_for_chunk(document, start_offset),
            page=document.page_number,
        )
        for chunk_number, (start_offset, end_offset) in enumerate(boundaries, start=1)
    )


def chunk_evidence_documents(
    documents: Sequence[EvidenceDocument],
    *,
    chunk_size: int = CHUNK_SIZE_CHARS,
    chunk_overlap: int = CHUNK_OVERLAP_CHARS,
) -> tuple[EvidenceChunk, ...]:
    """Chunk normalized evidence documents with stable source-level chunk IDs."""
    chunked_documents: list[EvidenceChunk] = []
    chunk_numbers_by_source: dict[str, int] = {}
    for document in documents:
        for chunk in chunk_evidence_document(
            document,
            chunk_size=chunk_size,
            chunk_overlap=chunk_overlap,
        ):
            next_chunk_number = chunk_numbers_by_source.get(chunk.source, 0) + 1
            chunk_numbers_by_source[chunk.source] = next_chunk_number
            chunked_documents.append(
                replace(
                    chunk,
                    chunk_id=(
                        f"{_chunk_id_prefix_for_source(chunk.source)}_"
                        f"{next_chunk_number:03d}"
                    ),
                    chunk_number=next_chunk_number,
                )
            )
    return tuple(chunked_documents)


def build_curated_evidence_chunks(
    evidence_dir: Path | None = None,
) -> tuple[EvidenceChunk, ...]:
    """Load, normalize, and chunk the curated evidence pack."""
    documents = load_curated_evidence_documents(evidence_dir)
    if not documents:
        raise ValueError("The curated evidence pack did not produce any loadable documents.")
    return chunk_evidence_documents(normalize_evidence_documents(documents))


def _chroma_metadata_for_chunk(chunk: EvidenceChunk) -> dict[str, str | int]:
    """Return storage-safe metadata for one chunk by dropping null Chroma fields."""
    return {
        key: value
        for key, value in chunk.metadata().items()
        if value is not None
    }


def persist_evidence_chunks(
    collection_handle: ChromaCollectionHandle,
    chunks: Sequence[EvidenceChunk],
) -> tuple[EvidenceChunk, ...]:
    """Upsert one deterministic chunk set into the persistent collection."""
    finalized_chunks = tuple(chunks)
    if not finalized_chunks:
        raise ValueError("At least one evidence chunk is required for persistence.")

    chunk_ids = [chunk.chunk_id for chunk in finalized_chunks]
    if any(chunk_id is None for chunk_id in chunk_ids):
        raise ValueError("All evidence chunks must have stable chunk IDs before persistence.")

    resolved_chunk_ids = [str(chunk_id) for chunk_id in chunk_ids]
    if len(set(resolved_chunk_ids)) != len(resolved_chunk_ids):
        raise ValueError("Evidence chunk IDs must be unique before persistence.")

    existing_collection_ids = set(collection_handle.collection.get()["ids"])
    stale_collection_ids = sorted(existing_collection_ids.difference(resolved_chunk_ids))
    if stale_collection_ids:
        collection_handle.collection.delete(ids=stale_collection_ids)

    collection_handle.collection.upsert(
        ids=resolved_chunk_ids,
        documents=[chunk.text for chunk in finalized_chunks],
        metadatas=[_chroma_metadata_for_chunk(chunk) for chunk in finalized_chunks],
    )
    return finalized_chunks


def persist_curated_evidence_chunks(
    *,
    collection_name: str = COLLECTION_NAME,
    persist_directory: Path | None = None,
    evidence_dir: Path | None = None,
    manifest_path: Path | None = None,
) -> tuple[EvidenceChunk, ...]:
    """Persist the canonical curated chunk set into the stable Chroma collection."""
    collection_handle = get_or_create_chroma_collection(
        collection_name=collection_name,
        persist_directory=persist_directory,
    )
    chunks = build_curated_evidence_chunks(evidence_dir)
    persisted_chunks = persist_evidence_chunks(collection_handle, chunks)
    record_indexed_workspace_state(
        collection_handle,
        workspace_hash=current_workspace_hash(manifest_path),
        chunk_count=len(persisted_chunks),
    )
    return persisted_chunks


def _persisted_index_status(
    *,
    collection_name: str = COLLECTION_NAME,
    persist_directory: Path | None = None,
    evidence_dir: Path | None = None,
    manifest_path: Path | None = None,
    index_action: str,
    reason: str,
) -> ChromaIndexStatus:
    """Persist the curated index and return the resulting ready status."""
    workspace_hash = current_workspace_hash(manifest_path)
    persisted_chunks = persist_curated_evidence_chunks(
        collection_name=collection_name,
        persist_directory=persist_directory,
        evidence_dir=evidence_dir,
        manifest_path=manifest_path,
    )
    collection_handle = get_existing_chroma_collection(
        collection_name=collection_name,
        persist_directory=persist_directory,
    )
    if collection_handle is None:
        raise RuntimeError(
            f"Persistent Chroma collection {collection_name!r} was not available after "
            "persisting the curated evidence chunks."
        )

    chunk_count = len(persisted_chunks)
    return ChromaIndexStatus(
        collection_name=collection_handle.collection_name,
        persist_directory=collection_handle.persist_directory,
        workspace_hash=workspace_hash,
        stored_workspace_hash=workspace_hash,
        stored_chunk_count=chunk_count,
        actual_chunk_count=chunk_count,
        index_action=index_action,
        ready=True,
        reason=reason,
        collection_handle=collection_handle,
    )


def create_curated_evidence_index(
    *,
    collection_name: str = COLLECTION_NAME,
    persist_directory: Path | None = None,
    evidence_dir: Path | None = None,
    manifest_path: Path | None = None,
) -> ChromaIndexStatus:
    """Create the canonical demo index when no reusable collection exists yet."""
    return _persisted_index_status(
        collection_name=collection_name,
        persist_directory=persist_directory,
        evidence_dir=evidence_dir,
        manifest_path=manifest_path,
        index_action=INDEX_ACTION_CREATED,
        reason="created",
    )


def delete_existing_chroma_collection(
    *,
    collection_name: str = COLLECTION_NAME,
    persist_directory: Path | None = None,
) -> bool:
    """Delete one existing persistent Chroma collection when it is present."""
    collection_handle = get_existing_chroma_collection(
        collection_name=collection_name,
        persist_directory=persist_directory,
    )
    if collection_handle is None:
        return False

    collection_handle.client.delete_collection(name=collection_handle.collection_name)
    return True


def rebuild_curated_evidence_index(
    *,
    collection_name: str = COLLECTION_NAME,
    persist_directory: Path | None = None,
    evidence_dir: Path | None = None,
    manifest_path: Path | None = None,
    index_action: str = INDEX_ACTION_REBUILT_CONTENT_CHANGE,
    reason: str = "workspace_hash_changed",
) -> ChromaIndexStatus:
    """Delete and recreate the canonical demo index for the current workspace."""
    delete_existing_chroma_collection(
        collection_name=collection_name,
        persist_directory=persist_directory,
    )
    return _persisted_index_status(
        collection_name=collection_name,
        persist_directory=persist_directory,
        evidence_dir=evidence_dir,
        manifest_path=manifest_path,
        index_action=index_action,
        reason=reason,
    )


def ensure_curated_evidence_index(
    *,
    collection_name: str = COLLECTION_NAME,
    persist_directory: Path | None = None,
    evidence_dir: Path | None = None,
    manifest_path: Path | None = None,
    force_rebuild: bool = False,
) -> ChromaIndexStatus:
    """Create, reuse, or rebuild the canonical demo index as needed."""
    reuse_status = evaluate_chroma_reuse(
        collection_name=collection_name,
        persist_directory=persist_directory,
        evidence_dir=evidence_dir,
        manifest_path=manifest_path,
    )
    if force_rebuild:
        if reuse_status.reason in BLOCKED_INDEX_REASONS:
            return _index_status_from_reuse_status(reuse_status)
        if reuse_status.reason == "collection_missing":
            return create_curated_evidence_index(
                collection_name=collection_name,
                persist_directory=persist_directory,
                evidence_dir=evidence_dir,
                manifest_path=manifest_path,
            )
        return rebuild_curated_evidence_index(
            collection_name=collection_name,
            persist_directory=persist_directory,
            evidence_dir=evidence_dir,
            manifest_path=manifest_path,
            index_action=INDEX_ACTION_REBUILT_CONTENT_CHANGE,
            reason="force_rebuild",
        )

    if reuse_status.reusable:
        return _index_status_from_reuse_status(reuse_status)

    if reuse_status.reason == "collection_missing":
        return create_curated_evidence_index(
            collection_name=collection_name,
            persist_directory=persist_directory,
            evidence_dir=evidence_dir,
            manifest_path=manifest_path,
        )

    if reuse_status.reason == "workspace_hash_mismatch":
        return rebuild_curated_evidence_index(
            collection_name=collection_name,
            persist_directory=persist_directory,
            evidence_dir=evidence_dir,
            manifest_path=manifest_path,
            index_action=INDEX_ACTION_REBUILT_CONTENT_CHANGE,
            reason="workspace_hash_changed",
        )

    if reuse_status.reason in AUTO_REBUILD_INTEGRITY_REASONS:
        return rebuild_curated_evidence_index(
            collection_name=collection_name,
            persist_directory=persist_directory,
            evidence_dir=evidence_dir,
            manifest_path=manifest_path,
            index_action=INDEX_ACTION_REBUILT_INTEGRITY,
            reason=reuse_status.reason,
        )

    return _index_status_from_reuse_status(reuse_status)


def _require_ready_collection_handle(
    index_status: ChromaIndexStatus,
) -> ChromaCollectionHandle:
    """Return the validated collection handle required for retrieval."""
    if not index_status.ready or index_status.collection_handle is None:
        raise ValueError(
            "The curated evidence index is not ready for retrieval. Run "
            "`ensure_curated_evidence_index()` and confirm it returned a ready "
            "status before retrieving evidence."
        )
    return index_status.collection_handle


def _question_identifier(row_like: Mapping[str, object]) -> str:
    """Return one best-effort row identifier for retrieval errors."""
    raw_question_id = row_like.get("question_id", row_like.get("Question ID", ""))
    if isinstance(raw_question_id, str) and raw_question_id.strip():
        return raw_question_id.strip()
    return "<unknown-question>"


def _load_dotenv_if_available() -> None:
    """Load repo-local .env values when python-dotenv is available."""
    try:
        from dotenv import load_dotenv
    except ModuleNotFoundError:
        return
    load_dotenv(REPO_ROOT / ".env", override=False)


def _openai_answer_client(openai_client: Any | None = None) -> Any:
    """Return the OpenAI client used for provider-backed answer generation."""
    if openai_client is not None:
        return openai_client

    _load_dotenv_if_available()
    openai_api_key = os.getenv("OPENAI_API_KEY", "").strip()
    if not openai_api_key:
        raise RuntimeError(
            "OPENAI_API_KEY is required for provider-backed answer generation. Set it "
            "in the shell or repo-local `.env` before calling the answer pipeline."
        )

    try:
        from openai import OpenAI
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            "openai is required for provider-backed answer generation. Install "
            "dependencies with `pip install -r requirements.txt`."
        ) from exc
    return OpenAI(api_key=openai_api_key)


def format_retrieved_chunk_for_prompt(chunk: RetrievedEvidenceChunk) -> str:
    """Render one retrieved chunk in the readable prompt shape from the plan."""
    metadata_parts = [f"source={chunk.source}"]
    if chunk.section:
        metadata_parts.append(f"section={chunk.section}")
    if chunk.page is not None:
        metadata_parts.append(f"page={chunk.page}")
    metadata_suffix = " ".join(metadata_parts)
    return f"[{chunk.chunk_id}] {metadata_suffix}\n{chunk.text}"


def build_answer_user_prompt(
    question_text: str,
    retrieved_chunks: Sequence[RetrievedEvidenceChunk],
) -> str:
    """Build the conservative user prompt with question text and retrieved evidence."""
    normalized_question_text = question_text.strip()
    if not normalized_question_text:
        raise ValueError("question_text must be a non-empty string.")

    evidence_section = "\n\n".join(
        format_retrieved_chunk_for_prompt(chunk) for chunk in retrieved_chunks
    )
    if not evidence_section:
        evidence_section = "<none provided>"

    return "\n".join(
        (
            "Question:",
            normalized_question_text,
            "",
            "Evidence chunks:",
            evidence_section,
        )
    )


def build_answer_prompt_messages(
    question_text: str,
    retrieved_chunks: Sequence[RetrievedEvidenceChunk],
) -> tuple[dict[str, str], ...]:
    """Return the exact system/user prompt messages for the answer-generation call."""
    return (
        {"role": "system", "content": CONSERVATIVE_ANSWER_SYSTEM_PROMPT},
        {
            "role": "user",
            "content": build_answer_user_prompt(question_text, retrieved_chunks),
        },
    )


def generate_answer_payload(
    question_text: str,
    retrieved_chunks: Sequence[RetrievedEvidenceChunk],
    *,
    model: str = DEFAULT_OPENAI_ANSWER_MODEL,
    openai_client: Any | None = None,
) -> dict[str, object]:
    """Call the provider-backed answer path and return the raw structured JSON payload."""
    normalized_model = model.strip()
    if not normalized_model:
        raise ValueError("model must be a non-empty string.")

    client = _openai_answer_client(openai_client)
    response = client.chat.completions.create(
        model=normalized_model,
        messages=list(build_answer_prompt_messages(question_text, retrieved_chunks)),
        response_format={
            "type": "json_schema",
            "json_schema": {
                "name": ANSWER_RESPONSE_SCHEMA_NAME,
                "schema": ANSWER_RESPONSE_JSON_SCHEMA,
                "strict": True,
            },
        },
    )

    choices = getattr(response, "choices", None)
    if not isinstance(choices, list) or not choices:
        raise RuntimeError("OpenAI answer generation returned no completion choices.")

    message = getattr(choices[0], "message", None)
    content = getattr(message, "content", None)
    if not isinstance(content, str) or not content.strip():
        raise RuntimeError("OpenAI answer generation returned an empty JSON payload.")

    try:
        payload = json.loads(content)
    except json.JSONDecodeError as exc:
        raise RuntimeError(
            "OpenAI answer generation returned malformed JSON for the answer payload."
        ) from exc

    if not isinstance(payload, dict):
        raise RuntimeError("OpenAI answer generation returned a non-object JSON payload.")
    return payload


def _reviewer_note_with_fallback(raw_value: object) -> str:
    """Return a short reviewer-oriented note, using a safe fallback when blank."""
    if isinstance(raw_value, str):
        normalized_value = raw_value.strip()
        if normalized_value:
            return normalized_value
    return FALLBACK_REVIEWER_NOTE


def validate_answer_payload(
    payload: Mapping[str, object],
    retrieved_chunks: Sequence[RetrievedEvidenceChunk],
) -> AnswerPayloadValidationResult:
    """Validate one raw answer payload and drop citations outside the retrieved set."""
    payload_keys = set(payload.keys())
    required_keys = set(ANSWER_OUTPUT_KEYS)
    if payload_keys != required_keys:
        return AnswerPayloadValidationResult(
            usable=False,
            outcome=ANSWER_PAYLOAD_OUTCOME_REJECTED,
            reviewer_note=FALLBACK_REVIEWER_NOTE,
            failure_reason="payload_shape_invalid",
        )

    answer = payload.get("answer")
    if not isinstance(answer, str) or not answer.strip():
        return AnswerPayloadValidationResult(
            usable=False,
            outcome=ANSWER_PAYLOAD_OUTCOME_REJECTED,
            reviewer_note=_reviewer_note_with_fallback(payload.get("reviewer_note")),
            failure_reason="answer_invalid",
        )
    normalized_answer = answer.strip()

    answer_type = payload.get("answer_type")
    if not isinstance(answer_type, str) or answer_type not in ANSWER_TYPES:
        return AnswerPayloadValidationResult(
            usable=False,
            outcome=ANSWER_PAYLOAD_OUTCOME_REJECTED,
            reviewer_note=_reviewer_note_with_fallback(payload.get("reviewer_note")),
            failure_reason="answer_type_invalid",
        )

    raw_citation_ids = payload.get("citation_ids")
    if not isinstance(raw_citation_ids, list):
        return AnswerPayloadValidationResult(
            usable=False,
            outcome=ANSWER_PAYLOAD_OUTCOME_REJECTED,
            reviewer_note=_reviewer_note_with_fallback(payload.get("reviewer_note")),
            failure_reason="citation_ids_invalid",
        )

    retrieved_chunk_ids = {chunk.chunk_id for chunk in retrieved_chunks}
    valid_citation_ids: list[str] = []
    seen_valid_citation_ids: set[str] = set()
    invalid_citation_ids: list[str] = []
    seen_invalid_citation_ids: set[str] = set()
    for raw_citation_id in raw_citation_ids:
        if not isinstance(raw_citation_id, str) or not raw_citation_id.strip():
            return AnswerPayloadValidationResult(
                usable=False,
                outcome=ANSWER_PAYLOAD_OUTCOME_REJECTED,
                reviewer_note=_reviewer_note_with_fallback(payload.get("reviewer_note")),
                failure_reason="citation_ids_invalid",
            )

        citation_id = raw_citation_id.strip()
        if citation_id in retrieved_chunk_ids:
            if citation_id not in seen_valid_citation_ids:
                seen_valid_citation_ids.add(citation_id)
                valid_citation_ids.append(citation_id)
            continue
        if citation_id not in seen_invalid_citation_ids:
            seen_invalid_citation_ids.add(citation_id)
            invalid_citation_ids.append(citation_id)

    reviewer_note = _reviewer_note_with_fallback(payload.get("reviewer_note"))
    if answer_type != ANSWER_TYPE_UNSUPPORTED and not valid_citation_ids:
        return AnswerPayloadValidationResult(
            usable=False,
            outcome=ANSWER_PAYLOAD_OUTCOME_REJECTED,
            reviewer_note=reviewer_note,
            answer=normalized_answer,
            answer_type=answer_type,
            citation_ids=(),
            invalid_citation_ids=tuple(invalid_citation_ids),
            failure_reason="no_valid_citations",
        )

    outcome = (
        ANSWER_PAYLOAD_OUTCOME_ACCEPTED_WITH_CITATION_IDS_REMOVED
        if invalid_citation_ids
        else ANSWER_PAYLOAD_OUTCOME_ACCEPTED
    )
    return AnswerPayloadValidationResult(
        usable=True,
        outcome=outcome,
        reviewer_note=reviewer_note,
        answer=normalized_answer,
        answer_type=answer_type,
        citation_ids=tuple(valid_citation_ids),
        invalid_citation_ids=tuple(invalid_citation_ids),
    )


def resolve_validated_citations(
    citation_ids: Sequence[str],
    retrieved_chunks: Sequence[RetrievedEvidenceChunk],
    *,
    max_visible: int = MAX_VISIBLE_CITATIONS,
) -> tuple[ResolvedEvidenceCitation, ...]:
    """Resolve validated citation ids into shared reviewer-facing evidence objects."""
    if max_visible <= 0:
        raise ValueError("max_visible must be a positive integer.")

    retrieved_chunks_by_id = {
        retrieved_chunk.chunk_id: retrieved_chunk for retrieved_chunk in retrieved_chunks
    }
    resolved_citations: list[ResolvedEvidenceCitation] = []
    for citation_id in citation_ids:
        if len(resolved_citations) >= max_visible:
            break

        retrieved_chunk = retrieved_chunks_by_id.get(citation_id)
        if retrieved_chunk is None:
            raise ValueError(
                "Validated citation id "
                f"{citation_id!r} was not present in the retrieved chunk set."
            )

        resolved_citations.append(
            ResolvedEvidenceCitation(
                chunk_id=retrieved_chunk.chunk_id,
                display_label=build_citation_display_label(
                    retrieved_chunk.source,
                    retrieved_chunk.section,
                    retrieved_chunk.page,
                ),
                snippet_text=retrieved_chunk.text,
                source=retrieved_chunk.source,
                source_path=retrieved_chunk.source_path,
                doc_type=retrieved_chunk.doc_type,
                section=retrieved_chunk.section,
                page=retrieved_chunk.page,
            )
        )

    return tuple(resolved_citations)


def _fail_closed_reviewer_note(failure_reason: str) -> str:
    """Return the short reviewer-facing rationale for one fail-closed reason."""
    return FAIL_CLOSED_REVIEWER_NOTE_BY_REASON.get(
        failure_reason,
        "Answer generation failed safely; review manually.",
    )


def _build_fail_closed_answer_result(
    failure_reason: str,
    *,
    retry_count: int,
) -> GeneratedAnswerResult:
    """Return the exact fail-closed answer package for one row-level failure."""
    confidence = score_answer_confidence(
        ANSWER_TYPE_UNSUPPORTED,
        valid_citation_count=0,
    )
    return GeneratedAnswerResult(
        answer=FAIL_CLOSED_ANSWER,
        answer_type=ANSWER_TYPE_UNSUPPORTED,
        citation_ids=(),
        citations=(),
        confidence_score=confidence.confidence_score,
        confidence_band=confidence.confidence_band,
        status=confidence.status,
        reviewer_note=_fail_closed_reviewer_note(failure_reason),
        failed_closed=True,
        failure_reason=failure_reason,
        retry_count=retry_count,
    )


def _is_retryable_generation_error(error: RuntimeError) -> bool:
    """Return whether one answer-generation runtime error is safe to retry once."""
    return str(error).startswith("OpenAI answer generation returned ")


def _should_retry_validation_failure(failure_reason: str | None) -> bool:
    """Return whether one validation failure should consume the single retry."""
    return failure_reason in RETRYABLE_VALIDATION_FAILURE_REASONS


def generate_answer_result(
    question_text: str,
    retrieved_chunks: Sequence[RetrievedEvidenceChunk],
    *,
    model: str = DEFAULT_OPENAI_ANSWER_MODEL,
    openai_client: Any | None = None,
) -> GeneratedAnswerResult:
    """Generate one answer result with fail-closed routing and at most one retry."""
    if not retrieved_chunks:
        return _build_fail_closed_answer_result(
            FAILURE_REASON_NO_RETRIEVAL,
            retry_count=0,
        )

    retry_count = 0
    while True:
        try:
            payload = generate_answer_payload(
                question_text,
                retrieved_chunks,
                model=model,
                openai_client=openai_client,
            )
        except RuntimeError as error:
            if _is_retryable_generation_error(error):
                if retry_count < MODEL_RETRY_LIMIT:
                    retry_count += 1
                    continue
                return _build_fail_closed_answer_result(
                    FAILURE_REASON_MODEL_OUTPUT_INVALID,
                    retry_count=retry_count,
                )
            raise

        validation_result = validate_answer_payload(payload, retrieved_chunks)
        if not validation_result.usable:
            failure_reason = validation_result.failure_reason or "payload_shape_invalid"
            if (
                _should_retry_validation_failure(failure_reason)
                and retry_count < MODEL_RETRY_LIMIT
            ):
                retry_count += 1
                continue
            return _build_fail_closed_answer_result(
                failure_reason,
                retry_count=retry_count,
            )

        if not validation_result.citation_ids and validation_result.invalid_citation_ids:
            return _build_fail_closed_answer_result(
                "no_valid_citations",
                retry_count=retry_count,
            )

        if validation_result.answer is None or validation_result.answer_type is None:
            raise RuntimeError(
                "validate_answer_payload returned a usable result without answer fields."
            )

        resolved_citations = resolve_validated_citations(
            validation_result.citation_ids,
            retrieved_chunks,
        )
        confidence = score_answer_confidence(
            validation_result.answer_type,
            valid_citation_count=len(validation_result.citation_ids),
        )
        return GeneratedAnswerResult(
            answer=validation_result.answer,
            answer_type=validation_result.answer_type,
            citation_ids=validation_result.citation_ids,
            citations=resolved_citations,
            confidence_score=confidence.confidence_score,
            confidence_band=confidence.confidence_band,
            status=confidence.status,
            reviewer_note=validation_result.reviewer_note,
            invalid_citation_ids=validation_result.invalid_citation_ids,
            retry_count=retry_count,
        )


def retrieve_evidence_chunks(
    question_text: str,
    *,
    index_status: ChromaIndexStatus,
    top_k: int = RETRIEVAL_TOP_K,
) -> tuple[RetrievedEvidenceChunk, ...]:
    """Retrieve up to five unique evidence chunks for one question string."""
    if top_k <= 0:
        raise ValueError("top_k must be a positive integer.")

    normalized_question_text = question_text.strip()
    if not normalized_question_text:
        raise ValueError("question_text must be a non-empty string.")

    collection_handle = _require_ready_collection_handle(index_status)
    query_limit = min(max(top_k * 2, top_k), index_status.actual_chunk_count)
    if query_limit <= 0:
        return ()

    payload = collection_handle.collection.query(
        query_texts=[normalized_question_text],
        n_results=query_limit,
        include=["documents", "metadatas", "distances"],
    )
    raw_ids = payload.get("ids")
    raw_documents = payload.get("documents")
    raw_metadatas = payload.get("metadatas")
    raw_distances = payload.get("distances")
    if not (
        isinstance(raw_ids, list)
        and len(raw_ids) == 1
        and isinstance(raw_documents, list)
        and len(raw_documents) == 1
        and isinstance(raw_metadatas, list)
        and len(raw_metadatas) == 1
    ):
        raise RuntimeError(
            "Chroma query returned an unexpected payload shape for evidence retrieval."
        )

    ids = raw_ids[0]
    documents = raw_documents[0]
    metadatas = raw_metadatas[0]
    if not (
        isinstance(ids, list) and isinstance(documents, list) and isinstance(metadatas, list)
    ):
        raise RuntimeError(
            "Chroma query returned a malformed retrieval payload for evidence retrieval."
        )

    if isinstance(raw_distances, list) and len(raw_distances) == 1 and isinstance(
        raw_distances[0], list
    ):
        distances: list[object | None] = list(raw_distances[0])
    else:
        distances = [None] * len(ids)

    if not (len(ids) == len(documents) == len(metadatas) == len(distances)):
        raise RuntimeError(
            "Chroma query returned mismatched retrieval result lengths for the current "
            "question."
        )

    retrieved_chunks: list[RetrievedEvidenceChunk] = []
    seen_chunk_ids: set[str] = set()
    for raw_id, raw_document, raw_metadata, raw_distance in zip(
        ids,
        documents,
        metadatas,
        distances,
    ):
        if not isinstance(raw_id, str) or not raw_id.strip():
            raise RuntimeError("Chroma query returned a retrieval candidate without an id.")
        if not isinstance(raw_document, str) or not raw_document.strip():
            raise RuntimeError(
                f"Chroma query returned an empty document for retrieval candidate {raw_id!r}."
            )
        if not isinstance(raw_metadata, dict):
            raise RuntimeError(
                f"Chroma query returned missing metadata for retrieval candidate {raw_id!r}."
            )

        chunk_id = raw_metadata.get("chunk_id")
        if not isinstance(chunk_id, str) or not chunk_id.strip():
            raise RuntimeError(
                f"Chroma query returned candidate {raw_id!r} without a valid logical chunk id."
            )
        if chunk_id in seen_chunk_ids:
            continue

        source_name = raw_metadata.get("source")
        if not isinstance(source_name, str) or not source_name.strip():
            raise RuntimeError(
                f"Chroma query returned candidate {chunk_id!r} without a valid source."
            )

        doc_type = raw_metadata.get("doc_type")
        if not isinstance(doc_type, str) or not doc_type.strip():
            raise RuntimeError(
                f"Chroma query returned candidate {chunk_id!r} without a valid doc_type."
            )

        section = raw_metadata.get("section")
        if section is not None and not isinstance(section, str):
            raise RuntimeError(
                f"Chroma query returned candidate {chunk_id!r} with a non-string section."
            )
        page = raw_metadata.get("page")
        if page is not None and not isinstance(page, int):
            raise RuntimeError(
                f"Chroma query returned candidate {chunk_id!r} with a non-integer page."
            )

        distance: float | None = None
        if isinstance(raw_distance, (int, float)):
            distance = float(raw_distance)

        seen_chunk_ids.add(chunk_id)
        retrieved_chunks.append(
            RetrievedEvidenceChunk(
                chunk_id=chunk_id,
                source=source_name,
                source_path=runtime_evidence_directory() / source_name,
                doc_type=doc_type,
                text=raw_document,
                rank=len(retrieved_chunks) + 1,
                distance=distance,
                section=section,
                page=page,
            )
        )
        if len(retrieved_chunks) >= top_k:
            break

    return tuple(retrieved_chunks)


def retrieve_evidence_chunks_for_row(
    row_like: Mapping[str, object],
    *,
    index_status: ChromaIndexStatus,
    top_k: int = RETRIEVAL_TOP_K,
) -> tuple[RetrievedEvidenceChunk, ...]:
    """Retrieve evidence using the canonical question text stored on one row."""
    raw_question_text = row_like.get("question", row_like.get("Question", ""))
    if not isinstance(raw_question_text, str) or not raw_question_text.strip():
        raise ValueError(
            "Question row "
            f"{_question_identifier(row_like)} is missing a usable question string."
        )
    return retrieve_evidence_chunks(
        raw_question_text,
        index_status=index_status,
        top_k=top_k,
    )


def _parse_evidence_display_value(value: str) -> list[str]:
    """Split the friendly workbook evidence cell into individual labels."""
    return [label for label in value.split("; ") if label]


def _visible_export_cell_value(
    row_like: Mapping[str, object],
    column_name: str,
) -> str:
    """Render one visible export cell from the canonical internal row contract."""
    def string_value(value: object) -> str:
        return "" if value is None else str(value)

    if column_name == "Question ID":
        return string_value(row_like["question_id"])
    if column_name == "Category":
        return string_value(row_like["category"])
    if column_name == "Question":
        return string_value(row_like["question"])
    if column_name == "Answer":
        return string_value(row_like["answer"])
    if column_name == "Evidence":
        evidence_labels = row_like.get("evidence_labels", ())
        if isinstance(evidence_labels, (list, tuple)):
            return build_evidence_display_value(
                string_value(label) for label in evidence_labels if string_value(label)
            )
        return string_value(row_like.get("Evidence", ""))
    if column_name == "Confidence":
        return string_value(row_like["confidence_band"])
    if column_name == "Status":
        return string_value(row_like["status"])
    if column_name == "Reviewer Notes":
        return string_value(row_like["reviewer_note"])
    raise KeyError(f"Unsupported visible export column {column_name!r}.")


def build_answered_questionnaire_workbook(
    questionnaire: RuntimeQuestionnaire,
) -> Workbook:
    """Create a fresh answered-questionnaire workbook from the visible row contract."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = QUESTION_SHEET_NAME
    worksheet.append(list(VISIBLE_EXPORT_COLUMNS))
    for row in questionnaire.rows:
        worksheet.append([
            _visible_export_cell_value(row, column_name)
            for column_name in VISIBLE_EXPORT_COLUMNS
        ])
    _apply_answered_questionnaire_worksheet_style(worksheet)
    return workbook


def _status_fill_name_for_status(status: object) -> str | None:
    if status == STATUS_READY_FOR_REVIEW:
        return READY_STATUS_FILL
    if status == STATUS_NEEDS_REVIEW:
        return REVIEW_STATUS_FILL
    return None


def _apply_answered_questionnaire_worksheet_style(worksheet: Any) -> None:
    """Apply the planned readability and review-triage styling to the export sheet."""
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    header_font = Font(bold=True)
    default_alignment = Alignment(vertical="top")
    wrapped_alignment = Alignment(vertical="top", wrap_text=True)
    status_fills = {
        fill_name: PatternFill(fill_type="solid", fgColor=rgb)
        for fill_name, rgb in STATUS_FILL_RGB_BY_NAME.items()
    }

    for column_index, column_name in enumerate(VISIBLE_EXPORT_COLUMNS, start=1):
        column_letter = get_column_letter(column_index)
        worksheet.column_dimensions[column_letter].width = ANSWERED_WORKBOOK_COLUMN_WIDTHS[
            column_name
        ]
        alignment = (
            wrapped_alignment
            if column_name in WRAPPED_EXPORT_COLUMNS
            else default_alignment
        )
        for row_index, cell in enumerate(worksheet[column_letter], start=1):
            cell.alignment = alignment
            if row_index == 1:
                cell.font = header_font
                continue
            if column_name != "Status":
                continue
            fill_name = _status_fill_name_for_status(cell.value)
            if fill_name is None:
                continue
            cell.fill = status_fills[fill_name]


def write_answered_questionnaire(
    questionnaire: RuntimeQuestionnaire,
    *,
    output_dir: Path | None = None,
) -> Path:
    """Write the answered-questionnaire workbook to the canonical outputs directory."""
    output_path = answered_questionnaire_output_path(output_dir)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = build_answered_questionnaire_workbook(questionnaire)
    try:
        workbook.save(output_path)
    finally:
        workbook.close()
    return output_path


def prepare_questionnaire_run(
    questionnaire: RuntimeQuestionnaire,
) -> RuntimeQuestionnaire:
    """Return one run-scoped questionnaire copy with cleared result fields."""
    run_rows: list[dict[str, object]] = []
    for row in questionnaire.rows:
        run_row = dict(row)
        run_row.update(
            {
                "Answer": "",
                "Evidence": "",
                "Confidence": "",
                "Status": "",
                "Reviewer Notes": "",
                **make_result_row_defaults(),
            }
        )
        run_rows.append(run_row)
    return replace(questionnaire, rows=run_rows)


def update_row_with_answer_result(
    row_like: Mapping[str, object],
    answer_result: GeneratedAnswerResult,
    *,
    index_action: str,
    run_id: str,
) -> dict[str, object]:
    """Merge one generated answer result into the shared internal row contract."""
    normalized_run_id = run_id.strip()
    if not normalized_run_id:
        raise ValueError("run_id must be a non-empty string.")

    evidence_labels = [
        citation.display_label for citation in answer_result.citations if citation.display_label
    ]
    reviewer_note = _reviewer_note_with_fallback(answer_result.reviewer_note)
    updated_row = dict(row_like)
    updated_row.update(
        {
            "Answer": answer_result.answer,
            "Evidence": build_evidence_display_value(evidence_labels),
            "Confidence": answer_result.confidence_band,
            "Status": answer_result.status,
            "Reviewer Notes": reviewer_note,
            "answer": answer_result.answer,
            "answer_type": answer_result.answer_type,
            "citation_ids": list(answer_result.citation_ids),
            "citations": list(answer_result.citations),
            "confidence_score": answer_result.confidence_score,
            "confidence_band": answer_result.confidence_band,
            "status": answer_result.status,
            "reviewer_note": reviewer_note,
            "evidence_labels": evidence_labels,
            "index_action": index_action,
            "run_id": normalized_run_id,
        }
    )
    return updated_row


def run_questionnaire_answer_pipeline(
    questionnaire: RuntimeQuestionnaire,
    *,
    index_status: ChromaIndexStatus,
    run_id: str,
    model: str = DEFAULT_OPENAI_ANSWER_MODEL,
    openai_client: Any | None = None,
    on_row_completed: Callable[[RuntimeQuestionnaire, int], None] | None = None,
) -> RuntimeQuestionnaire:
    """Fill one run-scoped questionnaire in order and expose incremental updates."""
    run_questionnaire = prepare_questionnaire_run(questionnaire)
    for row_index, row in enumerate(run_questionnaire.rows):
        retrieved_chunks = retrieve_evidence_chunks_for_row(
            row,
            index_status=index_status,
        )
        question_text = str(row["question"])
        answer_result = generate_answer_result(
            question_text,
            retrieved_chunks,
            model=model,
            openai_client=openai_client,
        )
        updated_row = update_row_with_answer_result(
            row,
            answer_result,
            index_action=index_status.index_action,
            run_id=run_id,
        )
        run_questionnaire.rows[row_index] = updated_row
        if on_row_completed is not None:
            on_row_completed(run_questionnaire, row_index)
    return run_questionnaire


def _build_runtime_question_row(
    *,
    question_id: str,
    category: str,
    question: str,
    visible_values: Mapping[str, str],
) -> dict[str, object]:
    """Assemble one canonical row with both visible and internal-only fields."""
    answer_value = visible_values["Answer"]
    evidence_value = visible_values["Evidence"]
    confidence_value = visible_values["Confidence"]
    status_value = visible_values["Status"]
    reviewer_notes_value = visible_values["Reviewer Notes"]

    row = {
        "Question ID": question_id,
        "Category": category,
        "Question": question,
        "Answer": answer_value,
        "Evidence": evidence_value,
        "Confidence": confidence_value,
        "Status": status_value,
        "Reviewer Notes": reviewer_notes_value,
        "question_id": question_id,
        "category": category,
        "question": question,
        **make_result_row_defaults(),
    }
    row["answer"] = answer_value
    row["confidence_band"] = confidence_value
    row["status"] = status_value
    row["reviewer_note"] = reviewer_notes_value
    row["evidence_labels"] = _parse_evidence_display_value(evidence_value)
    return row


def load_runtime_questionnaire(
    workbook_path: Path | None = None,
) -> RuntimeQuestionnaire:
    """Load the curated runtime workbook and add planned output columns in memory."""
    path = workbook_path or runtime_questionnaire_path()
    if not path.exists():
        raise FileNotFoundError(
            "Runtime questionnaire workbook is missing at "
            f"{path}. Run `python generate_demo_data.py` to restore the curated demo "
            "workspace before starting the pipeline."
        )

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if QUESTION_SHEET_NAME not in workbook.sheetnames:
            raise ValueError(
                f"Runtime questionnaire workbook {path} is missing the "
                f"`{QUESTION_SHEET_NAME}` sheet."
            )

        worksheet = workbook[QUESTION_SHEET_NAME]
        row_iter = worksheet.iter_rows(values_only=True)
        try:
            header_row = next(row_iter)
        except StopIteration as exc:
            raise ValueError(f"Runtime questionnaire workbook {path} is empty.") from exc

        header_values = [_normalize_workbook_text_cell(value) for value in header_row]
        if not any(header_values):
            raise ValueError(f"Runtime questionnaire workbook {path} is empty.")

        header_indexes = {
            column_name: index
            for index, column_name in enumerate(header_values)
            if column_name
        }
        missing_seed_columns = [
            column_name
            for column_name in SEED_QUESTION_COLUMNS
            if column_name not in header_indexes
        ]
        if missing_seed_columns:
            raise ValueError(
                "Runtime questionnaire workbook is missing required seed columns: "
                f"{', '.join(missing_seed_columns)}."
            )

        loaded_rows: list[dict[str, object]] = []
        for row_number, row_values in enumerate(row_iter, start=2):
            if not any(value is not None and str(value).strip() for value in row_values):
                continue

            normalized_values = [
                _normalize_workbook_text_cell(value) for value in row_values
            ]
            question_id = normalized_values[header_indexes["Question ID"]]
            category = normalized_values[header_indexes["Category"]]
            question = normalized_values[header_indexes["Question"]]
            if not question_id or not category or not question:
                raise ValueError(
                    "Runtime questionnaire workbook has a row with missing required "
                    f"seed values at Excel row {row_number}."
                )

            visible_values = {
                column_name: (
                    normalized_values[header_indexes[column_name]]
                    if column_name in header_indexes
                    else ""
                )
                for column_name in VISIBLE_OUTPUT_COLUMNS
            }
            loaded_rows.append(
                _build_runtime_question_row(
                    question_id=question_id,
                    category=category,
                    question=question,
                    visible_values=visible_values,
                )
            )

        if not loaded_rows:
            raise ValueError(
                f"Runtime questionnaire workbook {path} does not contain any question rows."
            )
    finally:
        workbook.close()

    return RuntimeQuestionnaire(
        workbook_path=path,
        visible_columns=VISIBLE_EXPORT_COLUMNS,
        rows=loaded_rows,
    )


def review_priority_sort_key(row_like: Mapping[str, object]) -> tuple[float, int]:
    """Sort review rows by ascending confidence and then questionnaire order."""
    question_id = str(row_like["question_id"])
    return (float(row_like["confidence_score"]), question_order_index(question_id))


def verification_command_by_name(command_name: str) -> VerificationCommand:
    """Return one canonical verification command by its stable name."""
    for command in CANONICAL_VERIFICATION_COMMANDS:
        if command.name == command_name:
            return command
    raise KeyError(f"Unknown verification command: {command_name}")


def verification_sequence_shell_commands(
    command_names: Sequence[str],
) -> tuple[str, ...]:
    """Render one named verification sequence as shell-ready commands."""
    return tuple(
        verification_command_by_name(command_name).shell_command()
        for command_name in command_names
    )


__all__ = [
    "ACCESS_CONTROL_POLICY_FILE_NAME",
    "ANSWER_OPENING_TOKENS",
    "ANSWER_OUTPUT_KEYS",
    "ANSWER_PAYLOAD_OUTCOME_ACCEPTED",
    "ANSWER_PAYLOAD_OUTCOME_ACCEPTED_WITH_CITATION_IDS_REMOVED",
    "ANSWER_PAYLOAD_OUTCOME_REJECTED",
    "ANSWER_RESPONSE_JSON_SCHEMA",
    "ANSWER_RESPONSE_SCHEMA_NAME",
    "ANSWER_TYPE_PARTIAL",
    "ANSWER_TYPE_SUPPORTED",
    "ANSWER_TYPE_UNSUPPORTED",
    "ANSWER_TYPES",
    "ANSWERED_QUESTIONNAIRE_FILE_NAME",
    "answered_questionnaire_output_path",
    "APP_SUBTITLE",
    "APP_TITLE",
    "AnswerConfidenceResult",
    "AnswerPayloadValidationResult",
    "BACKUP_RECOVERY_POLICY_FILE_NAME",
    "CANONICAL_VERIFICATION_COMMANDS",
    "CHROMA_DIR",
    "ChromaCollectionHandle",
    "ChromaIndexStatus",
    "ChromaReuseStatus",
    "CHUNK_OVERLAP_CHARS",
    "CHUNK_SIZE_CHARS",
    "COLLECTION_NAME",
    "CONFIDENCE_BAND_HIGH",
    "CONFIDENCE_BAND_LOW",
    "CONFIDENCE_BAND_MEDIUM",
    "CONSERVATIVE_ANSWER_SYSTEM_PROMPT",
    "CLOSEOUT_AUDIT_RULES",
    "CURATED_PDF_EVIDENCE_FILE_NAMES",
    "CURATED_TEXT_EVIDENCE_FILE_NAMES",
    "DATA_DIR",
    "DEFAULT_OPENAI_ANSWER_MODEL",
    "DEMO_MODE_LABEL",
    "DOCUMENT_TYPE_MARKDOWN",
    "DOCUMENT_TYPE_PDF",
    "DOCUMENT_TYPE_TEXT",
    "E2E_TESTS_DIR",
    "E2E_TEST_LOGS_DIR",
    "ENCRYPTION_POLICY_FILE_NAME",
    "EvidenceChunk",
    "EvidenceDocument",
    "EXPECTED_OUTCOMES_FIXTURE_PATH",
    "EXPECTED_EVIDENCE_FILE_NAMES",
    "EXPECTED_QUESTION_IDS",
    "FAIL_CLOSED_SCORE",
    "FAIL_CLOSED_ANSWER",
    "FALLBACK_REVIEWER_NOTE",
    "FAILURE_REASON_MODEL_OUTPUT_INVALID",
    "FAILURE_REASON_NO_RETRIEVAL",
    "FULL_LOCAL_VALIDATION_COMMAND_NAMES",
    "GeneratedAnswerResult",
    "HIGH_CONFIDENCE_THRESHOLD",
    "INCIDENT_RESPONSE_POLICY_FILE_NAME",
    "INDEX_ACTION_BLOCKED",
    "INDEX_ACTION_CREATED",
    "INDEX_ACTION_REBUILT_CONTENT_CHANGE",
    "INDEX_ACTION_REBUILT_INTEGRITY",
    "INDEX_ACTION_REUSED",
    "INDEX_FIXTURES_DIR",
    "LIVE_SMOKE_LOGS_DIR",
    "LOGGING_CONVENTIONS",
    "LOG_COMPONENTS",
    "LOG_COMPONENT_EXPORT",
    "LOG_COMPONENT_INDEXING",
    "LOG_COMPONENT_PIPELINE",
    "LOG_COMPONENT_SETUP",
    "LOG_COMPONENT_UI",
    "LOG_COMPONENT_VERIFICATION",
    "LogFieldSpec",
    "LOG_FIELD_SPECS",
    "LOG_FILE_EXTENSION",
    "LOG_FORMAT_NAME",
    "LOG_LEVELS",
    "LOG_OPTIONAL_FIELDS",
    "LOG_REQUIRED_FIELDS",
    "LOG_STATUSES",
    "LOG_STATUS_BLOCKED",
    "LOG_STATUS_COMPLETED",
    "LOG_STATUS_FAILED",
    "LOG_STATUS_RETRYING",
    "LOG_STATUS_SKIPPED",
    "LOG_STATUS_STARTED",
    "MAIN_RESULTS_TABLE_COLUMNS",
    "MANIFEST_FILE_NAME",
    "MAX_VISIBLE_CITATIONS",
    "MEDIUM_CONFIDENCE_THRESHOLD",
    "MODEL_RETRY_LIMIT",
    "NEEDS_REVIEW_FILE_NAME",
    "OPTIONAL_LIVE_VALIDATION_COMMAND_NAMES",
    "OUTPUTS_DIR",
    "PARTIAL_SCORE",
    "QUESTIONNAIRE_FILE_NAME",
    "QUESTION_SHEET_NAME",
    "QUICK_CONFIDENCE_COMMAND_NAMES",
    "READY_STATUS_FILL",
    "REPO_ROOT",
    "REQUIRED_ENV_VARS",
    "RETRIEVAL_TOP_K",
    "REVIEW_QUEUE_COLUMNS",
    "REVIEW_QUEUE_THRESHOLD",
    "REVIEW_STATUS_FILL",
    "REVIEW_SUMMARY_FILE_NAME",
    "ResolvedEvidenceCitation",
    "RetrievedEvidenceChunk",
    "RUNTIME_DIRECTORIES",
    "RUNTIME_EVIDENCE_DIR",
    "RUNTIME_QUESTIONNAIRES_DIR",
    "SEED_DATA_DIR",
    "SEED_EVIDENCE_DIR",
    "SEED_QUESTIONNAIRE_DIR",
    "SEED_QUESTION_COLUMNS",
    "SEED_TO_RUNTIME_PATHS",
    "SOC2_SUMMARY_FILE_NAME",
    "STUBBED_OPENAI_FIXTURES_DIR",
    "STATUS_NEEDS_REVIEW",
    "STATUS_READY_FOR_REVIEW",
    "SUMMARY_CARD_LABELS",
    "SUPPORTED_WITH_ONE_CITATION_SCORE",
    "SUPPORTED_WITH_TWO_PLUS_CITATIONS_SCORE",
    "TESTS_DIR",
    "TEST_FIXTURES_DIR",
    "TestSeam",
    "TEST_SEAMS",
    "TEXT_EVIDENCE_SUFFIXES",
    "UI_TESTS_DIR",
    "UI_TEST_LOGS_DIR",
    "UNIT_TESTS_DIR",
    "UNIT_TEST_LOGS_DIR",
    "UNSUPPORTED_SCORE",
    "VERIFICATION_ARTIFACTS_DIR",
    "VerificationCommand",
    "VISIBLE_EXPORT_COLUMNS",
    "VISIBLE_OUTPUT_COLUMNS",
    "WORKSPACE_HASH_DIRECTORIES",
    "WORKSPACE_FIXTURES_DIR",
    "build_answer_prompt_messages",
    "build_answered_questionnaire_workbook",
    "build_answer_user_prompt",
    "build_citation_display_label",
    "build_curated_evidence_chunks",
    "build_evidence_display_value",
    "chroma_persist_directory",
    "chunk_evidence_document",
    "chunk_evidence_documents",
    "confidence_band_for_score",
    "current_workspace_hash",
    "create_curated_evidence_index",
    "delete_existing_chroma_collection",
    "ensure_curated_evidence_index",
    "evaluate_chroma_reuse",
    "format_retrieved_chunk_for_prompt",
    "generate_answer_payload",
    "generate_answer_result",
    "get_existing_chroma_collection",
    "get_or_create_chroma_collection",
    "get_or_create_demo_chroma_collection",
    "load_curated_pdf_evidence_documents",
    "load_curated_evidence_documents",
    "load_curated_text_evidence_documents",
    "load_pdf_evidence_pages",
    "load_runtime_questionnaire",
    "load_text_evidence_document",
    "make_result_row_defaults",
    "normalize_evidence_document",
    "normalize_evidence_documents",
    "normalize_evidence_text",
    "persist_curated_evidence_chunks",
    "persist_evidence_chunks",
    "prepare_questionnaire_run",
    "question_order_index",
    "rebuild_curated_evidence_index",
    "record_indexed_workspace_state",
    "retrieve_evidence_chunks",
    "retrieve_evidence_chunks_for_row",
    "resolve_validated_citations",
    "review_priority_sort_key",
    "review_status_for_score",
    "run_questionnaire_answer_pipeline",
    "RuntimeQuestionnaire",
    "runtime_evidence_directory",
    "runtime_manifest_path",
    "runtime_questionnaire_path",
    "score_answer_confidence",
    "update_row_with_answer_result",
    "validate_answer_payload",
    "verification_command_by_name",
    "verification_sequence_shell_commands",
    "write_answered_questionnaire",
]
