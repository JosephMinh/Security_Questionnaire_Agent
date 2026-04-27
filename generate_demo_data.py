"""Demo workspace setup and reset entrypoint for the Security Questionnaire Agent."""

from __future__ import annotations

import argparse
from dataclasses import dataclass
import hashlib
import importlib.util
import json
from pathlib import Path
import shutil

from openpyxl import load_workbook

from rag import (
    CHROMA_DIR,
    DATA_DIR,
    ENCRYPTION_POLICY_FILE_NAME,
    EXPECTED_EVIDENCE_FILE_NAMES,
    EXPECTED_QUESTION_IDS,
    MANIFEST_FILE_NAME,
    OUTPUTS_DIR,
    QUESTIONNAIRE_FILE_NAME,
    QUESTION_SHEET_NAME,
    RUNTIME_DIRECTORIES,
    RUNTIME_EVIDENCE_DIR,
    RUNTIME_QUESTIONNAIRES_DIR,
    SEED_QUESTION_COLUMNS,
    SEED_TO_RUNTIME_PATHS,
    SOC2_SUMMARY_FILE_NAME,
    WORKSPACE_HASH_DIRECTORIES,
)


@dataclass(frozen=True)
class WorkspaceCopyResult:
    """One copied seed asset and the runtime path it was written to."""

    source_path: Path
    runtime_path: Path


@dataclass(frozen=True)
class CleanupResult:
    """One runtime artifact removed during setup cleanup."""

    removed_path: Path


@dataclass(frozen=True)
class ManifestEntry:
    """Hash metadata for one runtime workspace file."""

    relative_path: str
    sha256: str
    size_bytes: int


@dataclass(frozen=True)
class ValidationIssue:
    """One actionable runtime workspace validation problem."""

    path: Path
    message: str
    recovery_hint: str

    def render(self) -> str:
        """Return one user-facing validation line."""
        return f"- {self.path}: {self.message} Recovery: {self.recovery_hint}"


class WorkspaceValidationError(RuntimeError):
    """Raised when the curated runtime workspace no longer matches the demo contract."""

    def __init__(self, issues: tuple[ValidationIssue, ...]):
        self.issues = issues
        super().__init__("\n".join(issue.render() for issue in issues))


PDF_EXPECTED_TEXT_SNIPPETS = (
    "SOC 2 Type II",
    "independent third-party audit firm",
    "Audit period:",
    "relevant security controls",
)


def ensure_runtime_directories() -> tuple[Path, ...]:
    """Create the planned runtime directories when they do not already exist."""
    for directory in RUNTIME_DIRECTORIES:
        directory.mkdir(parents=True, exist_ok=True)
    return RUNTIME_DIRECTORIES


def copy_seed_assets() -> tuple[WorkspaceCopyResult, ...]:
    """Copy the curated seed files into the runtime workspace."""
    copied_assets: list[WorkspaceCopyResult] = []
    for source_path, runtime_path in SEED_TO_RUNTIME_PATHS:
        runtime_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(source_path, runtime_path)
        copied_assets.append(
            WorkspaceCopyResult(
                source_path=source_path,
                runtime_path=runtime_path,
            )
        )
    return tuple(copied_assets)


def iter_runtime_files(root_directory: Path) -> tuple[Path, ...]:
    """Return stable, sorted runtime files under one workspace directory."""
    return tuple(
        path
        for path in sorted(root_directory.rglob("*"))
        if path.is_file() and path.name != ".gitkeep"
    )


def runtime_file_names(directory: Path) -> set[str]:
    """Return the non-gitkeep file names currently present in one runtime directory."""
    if not directory.exists():
        return set()
    return {
        path.name
        for path in directory.iterdir()
        if path.is_file() and path.name != ".gitkeep"
    }


def clear_directory_contents(directory: Path) -> tuple[CleanupResult, ...]:
    """Remove runtime artifacts from one directory while preserving the directory itself."""
    removed_artifacts: list[CleanupResult] = []
    for path in sorted(directory.rglob("*"), reverse=True):
        if path.name == ".gitkeep":
            continue
        if path.is_file() or path.is_symlink():
            path.unlink()
            removed_artifacts.append(CleanupResult(removed_path=path))
            continue
        if path.is_dir():
            path.rmdir()
    return tuple(removed_artifacts)


def clear_output_artifacts() -> tuple[CleanupResult, ...]:
    """Clear prior output files without removing the output directory itself."""
    return clear_directory_contents(OUTPUTS_DIR)


def reset_index_cache() -> tuple[CleanupResult, ...]:
    """Clear only the local Chroma cache directory when explicitly requested."""
    return clear_directory_contents(CHROMA_DIR)


def clear_runtime_workspace_artifacts() -> tuple[CleanupResult, ...]:
    """Clear the runtime questionnaire and evidence directories before re-copying seed data."""
    return (
        *clear_directory_contents(RUNTIME_QUESTIONNAIRES_DIR),
        *clear_directory_contents(RUNTIME_EVIDENCE_DIR),
    )


def file_sha256(path: Path) -> str:
    """Compute a stable SHA-256 digest for one file."""
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(65536), b""):
            digest.update(chunk)
    return digest.hexdigest()


def build_manifest_entries() -> tuple[ManifestEntry, ...]:
    """Hash the runtime questionnaire and evidence files for manifest output."""
    entries: list[ManifestEntry] = []
    for root_directory in WORKSPACE_HASH_DIRECTORIES:
        for path in iter_runtime_files(root_directory):
            entries.append(
                ManifestEntry(
                    relative_path=str(path.relative_to(DATA_DIR)),
                    sha256=file_sha256(path),
                    size_bytes=path.stat().st_size,
                )
            )
    return tuple(entries)


def manifest_path() -> Path:
    """Return the runtime manifest path."""
    return DATA_DIR / MANIFEST_FILE_NAME


def questionnaire_path() -> Path:
    """Return the curated runtime questionnaire path."""
    return RUNTIME_QUESTIONNAIRES_DIR / QUESTIONNAIRE_FILE_NAME


def expected_runtime_evidence_paths() -> tuple[Path, ...]:
    """Return the curated runtime evidence paths in canonical order."""
    return tuple(RUNTIME_EVIDENCE_DIR / file_name for file_name in EXPECTED_EVIDENCE_FILE_NAMES)


def validate_questionnaire_workbook() -> tuple[ValidationIssue, ...]:
    """Validate the runtime questionnaire workbook against the fixed demo contract."""
    path = questionnaire_path()
    recovery_hint = (
        "Rerun `python generate_demo_data.py` to restore the curated workbook from seed data."
    )
    issues: list[ValidationIssue] = []
    existing_file_names = runtime_file_names(RUNTIME_QUESTIONNAIRES_DIR)
    unexpected_file_names = sorted(existing_file_names - {QUESTIONNAIRE_FILE_NAME})
    for file_name in unexpected_file_names:
        issues.append(
            ValidationIssue(
                path=RUNTIME_QUESTIONNAIRES_DIR / file_name,
                message="Unexpected runtime questionnaire file present outside the curated demo set.",
                recovery_hint=(
                    "Reset the demo workspace so only the bundled questionnaire workbook remains."
                ),
            )
        )

    if not path.exists():
        issues.append(
            ValidationIssue(
                path=path,
                message="The curated runtime questionnaire workbook is missing.",
                recovery_hint=recovery_hint,
            )
        )
        return tuple(issues)

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # pragma: no cover - defensive contract handling
        issues.append(
            ValidationIssue(
                path=path,
                message=f"The questionnaire workbook could not be opened: {exc}",
                recovery_hint=recovery_hint,
            )
        )
        return tuple(issues)

    try:
        if workbook.sheetnames != [QUESTION_SHEET_NAME]:
            issues.append(
                ValidationIssue(
                    path=path,
                    message=(
                        f"Expected exactly one worksheet named `{QUESTION_SHEET_NAME}`, "
                        f"found {workbook.sheetnames}."
                    ),
                    recovery_hint=recovery_hint,
                )
            )
            return tuple(issues)

        worksheet = workbook[QUESTION_SHEET_NAME]
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            issues.append(
                ValidationIssue(
                    path=path,
                    message="The questionnaire workbook is empty.",
                    recovery_hint=recovery_hint,
                )
            )
            return tuple(issues)

        observed_columns = tuple("" if cell is None else str(cell) for cell in rows[0])
        if observed_columns != SEED_QUESTION_COLUMNS:
            issues.append(
                ValidationIssue(
                    path=path,
                    message=(
                        f"Expected source columns {SEED_QUESTION_COLUMNS}, "
                        f"found {observed_columns}."
                    ),
                    recovery_hint=(
                        "Replace the workbook with the curated seed copy or rerun "
                        "`python generate_demo_data.py`."
                    ),
                )
            )

        observed_question_ids: list[str] = []
        seen_question_ids: set[str] = set()
        for row_index, row in enumerate(rows[1:], start=2):
            if len(row) < len(SEED_QUESTION_COLUMNS):
                issues.append(
                    ValidationIssue(
                        path=path,
                        message=f"Row {row_index} does not contain all required source columns.",
                        recovery_hint=recovery_hint,
                    )
                )
                continue

            question_id = "" if row[0] is None else str(row[0]).strip()
            category = "" if row[1] is None else str(row[1]).strip()
            question_text = "" if row[2] is None else str(row[2]).strip()

            if not question_id or not category or not question_text:
                issues.append(
                    ValidationIssue(
                        path=path,
                        message=(
                            f"Row {row_index} must populate `Question ID`, `Category`, and `Question`."
                        ),
                        recovery_hint=recovery_hint,
                    )
                )
                continue

            if question_id in seen_question_ids:
                issues.append(
                    ValidationIssue(
                        path=path,
                        message=f"Question ID `{question_id}` is duplicated in the workbook.",
                        recovery_hint=(
                            "Restore the canonical workbook from seed data so each demo question "
                            "appears exactly once."
                        ),
                    )
                )
                continue

            seen_question_ids.add(question_id)
            observed_question_ids.append(question_id)

        if observed_question_ids != list(EXPECTED_QUESTION_IDS):
            issues.append(
                ValidationIssue(
                    path=path,
                    message=(
                        f"Expected question IDs in order {list(EXPECTED_QUESTION_IDS)}, "
                        f"found {observed_question_ids}."
                    ),
                    recovery_hint=(
                        "Rerun `python generate_demo_data.py` to restore the canonical 22-question "
                        "demo workbook."
                    ),
                )
            )
    finally:
        workbook.close()

    return tuple(issues)


def validate_soc2_summary_pdf(path: Path) -> tuple[ValidationIssue, ...]:
    """Validate that the bundled SOC 2 PDF still looks like the intended text-based artifact."""
    recovery_hint = (
        "Regenerate the curated evidence workspace or replace the PDF with the bundled seed copy."
    )
    if not path.exists():
        return (
            ValidationIssue(
                path=path,
                message="The bundled SOC 2 summary PDF is missing.",
                recovery_hint=recovery_hint,
            ),
        )

    raw_bytes = path.read_bytes()
    if not raw_bytes:
        return (
            ValidationIssue(
                path=path,
                message="The bundled SOC 2 summary PDF is empty.",
                recovery_hint=recovery_hint,
            ),
        )
    if not raw_bytes.startswith(b"%PDF-"):
        return (
            ValidationIssue(
                path=path,
                message="The bundled SOC 2 summary does not have a valid PDF header.",
                recovery_hint=recovery_hint,
            ),
        )

    visible_text = raw_bytes.decode("latin-1", errors="ignore")
    missing_snippets = [
        snippet for snippet in PDF_EXPECTED_TEXT_SNIPPETS if snippet not in visible_text
    ]
    issues: list[ValidationIssue] = []
    if missing_snippets:
        issues.append(
            ValidationIssue(
                path=path,
                message=(
                    "The bundled SOC 2 PDF is missing required visible text claims: "
                    + ", ".join(missing_snippets)
                ),
                recovery_hint=recovery_hint,
            )
        )

    if importlib.util.find_spec("pypdf") is not None:
        try:
            from pypdf import PdfReader  # type: ignore

            extracted_text = "".join(page.extract_text() or "" for page in PdfReader(path).pages)
        except Exception as exc:  # pragma: no cover - depends on optional runtime dependency
            issues.append(
                ValidationIssue(
                    path=path,
                    message=f"The bundled SOC 2 PDF could not be parsed by pypdf: {exc}",
                    recovery_hint=recovery_hint,
                )
            )
        else:
            if not extracted_text.strip():
                issues.append(
                    ValidationIssue(
                        path=path,
                        message="The bundled SOC 2 PDF did not yield extractable text in pypdf.",
                        recovery_hint=recovery_hint,
                    )
                )

    return tuple(issues)


def validate_runtime_evidence_files() -> tuple[ValidationIssue, ...]:
    """Validate the curated runtime evidence file set against the fixed demo contract."""
    issues: list[ValidationIssue] = []
    recovery_hint = (
        "Rerun `python generate_demo_data.py` to repopulate the curated evidence workspace."
    )
    existing_file_names = runtime_file_names(RUNTIME_EVIDENCE_DIR)
    expected_file_names = set(EXPECTED_EVIDENCE_FILE_NAMES)

    missing_file_names = sorted(expected_file_names - existing_file_names)
    unexpected_file_names = sorted(existing_file_names - expected_file_names)

    for file_name in missing_file_names:
        issues.append(
            ValidationIssue(
                path=RUNTIME_EVIDENCE_DIR / file_name,
                message="The curated runtime evidence file is missing.",
                recovery_hint=recovery_hint,
            )
        )
    for file_name in unexpected_file_names:
        issues.append(
            ValidationIssue(
                path=RUNTIME_EVIDENCE_DIR / file_name,
                message="Unexpected runtime evidence file present outside the curated demo set.",
                recovery_hint="Reset the demo workspace so only the bundled evidence files remain.",
            )
        )

    for path in expected_runtime_evidence_paths():
        if not path.exists():
            continue
        if path.stat().st_size == 0:
            issues.append(
                ValidationIssue(
                    path=path,
                    message="The curated runtime evidence file is empty.",
                    recovery_hint=recovery_hint,
                )
            )
            continue
        if path.name == SOC2_SUMMARY_FILE_NAME:
            issues.extend(validate_soc2_summary_pdf(path))

    return tuple(issues)


def validate_runtime_workspace() -> None:
    """Fail fast if the curated runtime workspace no longer matches the demo contract."""
    issues = (
        *validate_questionnaire_workbook(),
        *validate_runtime_evidence_files(),
    )
    if issues:
        raise WorkspaceValidationError(tuple(issues))


def write_workspace_manifest() -> Path:
    """Write the current runtime workspace manifest with file-level hashes."""
    entries = build_manifest_entries()
    combined_digest = hashlib.sha256()
    for entry in entries:
        combined_digest.update(entry.relative_path.encode("utf-8"))
        combined_digest.update(b"\0")
        combined_digest.update(entry.sha256.encode("utf-8"))
        combined_digest.update(b"\0")

    payload = {
        "manifest_version": 1,
        "workspace_hash": combined_digest.hexdigest(),
        "files": [
            {
                "relative_path": entry.relative_path,
                "sha256": entry.sha256,
                "size_bytes": entry.size_bytes,
            }
            for entry in entries
        ],
    }
    path = manifest_path()
    path.write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    return path


def parse_args() -> argparse.Namespace:
    """Parse CLI flags for the setup script."""
    parser = argparse.ArgumentParser(
        description="Prepare the demo workspace from the curated seed assets."
    )
    parser.add_argument(
        "--reset-index",
        action="store_true",
        help="Clear only the local Chroma index cache before the next build.",
    )
    return parser.parse_args()


def prepare_demo_workspace(*, reset_index: bool = False) -> tuple[WorkspaceCopyResult, ...]:
    """Create the runtime workspace, clean it safely, and populate it with seed assets."""
    ensure_runtime_directories()
    clear_runtime_workspace_artifacts()
    copied_assets = copy_seed_assets()
    clear_output_artifacts()
    if reset_index:
        reset_index_cache()
    validate_runtime_workspace()
    write_workspace_manifest()
    return copied_assets


def main() -> int:
    """Run the setup path used by the UI's Load Demo Workspace action."""
    args = parse_args()
    try:
        copied_assets = prepare_demo_workspace(reset_index=args.reset_index)
    except WorkspaceValidationError as exc:
        print("Workspace validation failed.")
        for issue in exc.issues:
            print(issue.render())
        return 1
    print("Prepared demo workspace.")
    print(f"Workspace manifest: {manifest_path()}")
    print(f"Index reset requested: {args.reset_index}")
    for asset in copied_assets:
        print(f"{asset.source_path} -> {asset.runtime_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())


__all__ = [
    "ENCRYPTION_POLICY_FILE_NAME",
    "MANIFEST_FILE_NAME",
    "RUNTIME_DIRECTORIES",
    "SEED_TO_RUNTIME_PATHS",
    "CleanupResult",
    "ManifestEntry",
    "ValidationIssue",
    "WorkspaceCopyResult",
    "build_manifest_entries",
    "clear_directory_contents",
    "clear_output_artifacts",
    "copy_seed_assets",
    "ensure_runtime_directories",
    "file_sha256",
    "main",
    "manifest_path",
    "parse_args",
    "prepare_demo_workspace",
    "reset_index_cache",
    "validate_questionnaire_workbook",
    "validate_runtime_evidence_files",
    "validate_runtime_workspace",
    "write_workspace_manifest",
]
